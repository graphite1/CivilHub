from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


DEFAULT_MAPPING_SHEET = "セル対応表"


def load_ledger_cell_mappings(mapping_path: Path, sheet_name: str = DEFAULT_MAPPING_SHEET) -> list[dict[str, str]]:
    if not mapping_path.exists():
        raise FileNotFoundError(f"セル対応表が見つかりません: {mapping_path}")

    workbook = load_workbook(mapping_path, data_only=True)
    try:
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        header_index = {str(value): index for index, value in enumerate(headers) if value is not None}
        required = ["シート名", "代表セル", "アプリ項目名候補", "ユーザー確認"]
        missing = [name for name in required if name not in header_index]
        if missing:
            raise ValueError(f"セル対応表に必要な列がありません: {', '.join(missing)}")

        mappings: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            confirmation = row[header_index["ユーザー確認"]]
            if confirmation != "OK":
                continue
            source_sheet_name = row[header_index["シート名"]]
            cell = row[header_index["代表セル"]]
            field = row[header_index["アプリ項目名候補"]]
            if not source_sheet_name or not cell or not field:
                continue
            mappings.append(
                {
                    "mapping_no": _optional_row_value(row, header_index, "No", ""),
                    "sheet_name": str(source_sheet_name),
                    "cell": str(cell),
                    "cell_range": _optional_row_value(row, header_index, "セル/結合範囲", str(cell)),
                    "item_name": _optional_row_value(row, header_index, "項目名候補", str(field)),
                    "field": str(field),
                }
            )
        return mappings
    finally:
        workbook.close()


def _optional_row_value(row: tuple[object, ...], header_index: dict[str, int], key: str, default: str) -> str:
    if key not in header_index:
        return default
    value = row[header_index[key]]
    return str(value) if value else default
