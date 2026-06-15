from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from .ledger_mapping import DEFAULT_MAPPING_SHEET, load_ledger_cell_mappings


def parse_wareki_free_date(value: str) -> str:
    match = re.search(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", value)
    if not match:
        return value.strip()
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def clean_ledger_phone(value: str) -> str:
    cleaned = value.strip()
    cleaned = cleaned.replace("（TEL", "").replace("(TEL", "")
    cleaned = cleaned.replace("）", "").replace(")", "")
    return cleaned.strip()


def clean_ledger_address(value: str) -> str:
    cleaned = value.strip()
    cleaned = re.sub(r"^〒\s*", "", cleaned)
    return cleaned.strip()


def load_ledger_project_info(
    xlsx_path: Path,
    mapping_path: Path,
    mapping_sheet: str = DEFAULT_MAPPING_SHEET,
) -> dict[str, object] | None:
    mapped_items = read_ledger_mapped_items(xlsx_path, mapping_path, mapping_sheet)
    values = ledger_items_to_values(mapped_items)

    project_name = values.get("project.basic.name", "")
    if not project_name:
        return None

    return {
        "name": project_name,
        "client_name": values.get("client.name", ""),
        "start_date": parse_wareki_free_date(values.get("project.contract.period_start", "").replace("自", "").strip()),
        "end_date": parse_wareki_free_date(values.get("project.contract.period_end", "").replace("至", "").strip()),
        "site_agent": values.get("prime_contractor.engineers.site_agent_name", ""),
        "managing_engineer": values.get("prime_contractor.engineers.chief_engineer_name", ""),
        "chief_engineer": values.get("prime_contractor.engineers.chief_engineer_name", ""),
        "source_path": str(xlsx_path),
        "source_sheet": values.get("__sheet_name", ""),
        "mapped_items": mapped_items,
    }


def read_ledger_mapped_values(
    xlsx_path: Path,
    mapping_path: Path,
    mapping_sheet: str = DEFAULT_MAPPING_SHEET,
) -> dict[str, str]:
    return ledger_items_to_values(read_ledger_mapped_items(xlsx_path, mapping_path, mapping_sheet))


def ledger_items_to_values(mapped_items: list[dict[str, str]]) -> dict[str, str]:
    values: dict[str, str] = {}
    for item in mapped_items:
        if item["field"] == "__sheet_name":
            values["__sheet_name"] = item["actual_sheet_name"]
            continue
        if not item["value"]:
            continue
        values[item["field"]] = item["value"]
    return values


def read_ledger_mapped_items(
    xlsx_path: Path,
    mapping_path: Path,
    mapping_sheet: str = DEFAULT_MAPPING_SHEET,
) -> list[dict[str, str]]:
    mappings = load_ledger_cell_mappings(mapping_path, mapping_sheet)
    workbook = load_workbook(xlsx_path, data_only=True)
    try:
        source_sheet_name = workbook.sheetnames[0]
        source_sheet = workbook[source_sheet_name]
        mapped_items: list[dict[str, str]] = [
            {
                "sheet_name": source_sheet_name,
                "actual_sheet_name": source_sheet_name,
                "cell": "",
                "cell_range": "",
                "item_name": "読取シート",
                "field": "__sheet_name",
                "value": source_sheet_name,
            }
        ]
        for mapping in mappings:
            sheet_name = mapping["sheet_name"]
            if sheet_name in workbook.sheetnames:
                actual_sheet_name = sheet_name
                sheet = workbook[sheet_name]
            else:
                actual_sheet_name = source_sheet_name
                sheet = source_sheet
            value = sheet[mapping["cell"]].value
            mapped_items.append(
                {
                    "sheet_name": mapping["sheet_name"],
                    "actual_sheet_name": actual_sheet_name,
                    "cell": mapping["cell"],
                    "cell_range": mapping["cell_range"],
                    "item_name": mapping["item_name"],
                    "field": mapping["field"],
                    "value": "" if value in (None, "") else str(value).strip(),
                }
            )
        return mapped_items
    finally:
        workbook.close()
