from __future__ import annotations

import csv
import os
import re
import shutil
import sqlite3
import subprocess
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tkinter import BOTH, END, LEFT, RIGHT, VERTICAL, X, Y, filedialog, messagebox, simpledialog, StringVar, Tk, Toplevel
import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook


APP_TITLE = "CIVIL HUB"
APP_SUBTITLE = "施工体制台帳作成・添付書類チェックシステム"
DB_PATH = Path(__file__).with_name("civilhub.db")
ATTACHMENTS_ROOT = Path(__file__).with_name("storage") / "attachments"
HANDOFF_ROOT = Path(__file__).with_name("handoff")
EXPORTS_ROOT = Path(__file__).with_name("exports")
LEDGER_MASTER_PATH = HANDOFF_ROOT / "施工体制台帳_マスター.xlsx"
LEDGER_CELL_MAPPING_PATH = HANDOFF_ROOT / "施工体制台帳_セル対応表.xlsx"
LEDGER_CELL_MAPPING_SHEET = "セル対応表"
DOCUMENT_TEMPLATES = [
    "建設業許可証",
    "主任技術者資格証",
    "主任技術者の雇用確認資料",
    "社会保険加入確認資料",
    "労働保険関係資料",
    "安全衛生責任者選任書",
    "作業員名簿",
    "外国人就労関係書類",
    "契約書または注文請書",
]
DOCUMENT_STATUSES = ["未確認", "不足", "添付済み", "期限切れ", "不要"]
LEVEL_LABELS = {0: "元請", 1: "一次下請", 2: "二次下請", 3: "三次下請以降"}
LEDGER_DOCUMENT_KIND = "施工体制台帳"
RECONTRACT_DOCUMENT_KIND = "再下請負通知書"
STATUS_COLORS = {
    "添付済み": "#1f7a38",
    "不足": "#9f2d2d",
    "期限切れ": "#9f2d2d",
    "未確認": "#5f6368",
    "不要": "#607d8b",
}


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def sanitize_filename(value: str) -> str:
    forbidden = '<>:"/\\|?*'
    cleaned = "".join("_" if ch in forbidden else ch for ch in value.strip())
    return cleaned or "document"


def provisional_construction_no(project_name: str) -> str:
    safe_name = sanitize_filename(project_name).replace(" ", "_")
    return f"CFG_{safe_name}"[:80]


def normalize_label(value: str) -> str:
    return value.replace("\n", "").replace("\r", "").replace("\u3000", "").replace(" ", "").strip()


def parse_wareki_free_date(value: str) -> str:
    match = re.search(r"(\d{4})年\s*(\d{1,2})月\s*(\d{1,2})日", value)
    if not match:
        return value.strip()
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def format_ledger_date(prefix: str, value: str) -> str:
    if not value:
        return ""
    match = re.match(r"^\s*(\d{4})-(\d{1,2})-(\d{1,2})\s*$", value)
    if match:
        year, month, day = match.groups()
        return f" {prefix} {int(year)}年{int(month)}月{int(day)}日"
    return value


def clean_ledger_phone(value: str) -> str:
    cleaned = value.strip()
    cleaned = cleaned.replace("（TEL", "").replace("(TEL", "")
    cleaned = cleaned.replace("）", "").replace(")", "")
    return cleaned.strip()


def clean_ledger_address(value: str) -> str:
    cleaned = value.strip()
    cleaned = re.sub(r"^〒\s*", "", cleaned)
    return cleaned.strip()


def format_ledger_phone(value: str) -> str:
    if not value:
        return ""
    return f"（TEL　{value}）"


def build_ledger_export_path(project_name: str, company_name: str | None = None) -> Path:
    EXPORTS_ROOT.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    parts = ["施工体制台帳", sanitize_filename(project_name)]
    if company_name:
        parts.append(sanitize_filename(company_name))
    parts.append(timestamp)
    return EXPORTS_ROOT / ("_".join(parts) + ".xlsx")


def load_ledger_cell_mappings() -> list[dict[str, str]]:
    if not LEDGER_CELL_MAPPING_PATH.exists():
        raise FileNotFoundError(f"セル対応表が見つかりません: {LEDGER_CELL_MAPPING_PATH}")

    workbook = load_workbook(LEDGER_CELL_MAPPING_PATH, data_only=True)
    try:
        sheet = workbook[LEDGER_CELL_MAPPING_SHEET]
        headers = [cell.value for cell in sheet[1]]
        header_index = {str(value): index for index, value in enumerate(headers) if value is not None}
        required = ["シート名", "代表セル", "アプリ項目名候補", "ユーザー確認"]
        missing = [name for name in required if name not in header_index]
        if missing:
            raise ValueError(f"セル対応表に必要な列がありません: {', '.join(missing)}")

        mappings: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            confirmation = row[header_index["ユーザー確認"]]
            if confirmation != "OK":
                continue
            sheet_name = row[header_index["シート名"]]
            cell = row[header_index["代表セル"]]
            field = row[header_index["アプリ項目名候補"]]
            if not sheet_name or not cell or not field:
                continue
            mappings.append(
                {
                    "sheet_name": str(sheet_name),
                    "cell": str(cell),
                    "cell_range": str(row[header_index["セル/結合範囲"]]) if "セル/結合範囲" in header_index and row[header_index["セル/結合範囲"]] else str(cell),
                    "item_name": str(row[header_index["項目名候補"]]) if "項目名候補" in header_index and row[header_index["項目名候補"]] else str(field),
                    "field": str(field),
                }
            )
        return mappings
    finally:
        workbook.close()


def build_reflected_copy_path(source_path: Path) -> Path:
    base = source_path.with_name(f"{source_path.stem}_civilhub{source_path.suffix}")
    if not base.exists():
        return base
    version = 2
    while True:
        candidate = source_path.with_name(f"{source_path.stem}_civilhub_v{version}{source_path.suffix}")
        if not candidate.exists():
            return candidate
        version += 1


def relation_document_kind(parent_level: int, child_level: int) -> str:
    if parent_level == 0 and child_level == 1:
        return LEDGER_DOCUMENT_KIND
    return RECONTRACT_DOCUMENT_KIND


def open_path(path: Path) -> None:
    if sys.platform.startswith("win"):
        os.startfile(str(path))  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.run(["open", str(path)], check=False)
    else:
        subprocess.run(["xdg-open", str(path)], check=False)


class Database:
    def __init__(self, db_path: Path) -> None:
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row
        self.initialize()

    def initialize(self) -> None:
        self.conn.executescript(
            """
            PRAGMA foreign_keys = ON;

            CREATE TABLE IF NOT EXISTS projects (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                construction_no TEXT NOT NULL,
                client_name TEXT NOT NULL,
                start_date TEXT,
                end_date TEXT,
                site_agent TEXT,
                managing_engineer TEXT,
                chief_engineer TEXT,
                base_folder TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS companies (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                parent_company_id INTEGER,
                level INTEGER NOT NULL,
                name TEXT NOT NULL,
                kana TEXT,
                representative TEXT,
                address TEXT,
                phone TEXT,
                license_no TEXT,
                license_expiry TEXT,
                work_type TEXT,
                contract_date TEXT,
                planned_start_date TEXT,
                planned_end_date TEXT,
                chief_engineer_name TEXT,
                chief_engineer_license TEXT,
                safety_manager TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE,
                FOREIGN KEY(parent_company_id) REFERENCES companies(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS required_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id INTEGER NOT NULL,
                document_name TEXT NOT NULL,
                required INTEGER NOT NULL DEFAULT 1,
                status TEXT NOT NULL DEFAULT '未確認',
                expiry_date TEXT,
                note TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(company_id) REFERENCES companies(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                required_document_id INTEGER NOT NULL,
                original_path TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                stored_filename TEXT NOT NULL,
                file_type TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(required_document_id) REFERENCES required_documents(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS project_imports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                import_kind TEXT NOT NULL,
                source_path TEXT NOT NULL,
                source_sheet TEXT,
                imported_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS company_relation_documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL,
                parent_company_id INTEGER NOT NULL,
                child_company_id INTEGER NOT NULL,
                document_kind TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT '未確認',
                source_path TEXT,
                source_sheet TEXT,
                reflected_path TEXT,
                note TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE,
                FOREIGN KEY(parent_company_id) REFERENCES companies(id) ON DELETE CASCADE,
                FOREIGN KEY(child_company_id) REFERENCES companies(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS relation_document_attachments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                relation_document_id INTEGER NOT NULL,
                original_path TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                original_filename TEXT NOT NULL,
                stored_filename TEXT NOT NULL,
                file_type TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(relation_document_id) REFERENCES company_relation_documents(id) ON DELETE CASCADE
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_company_relation_documents_pair_kind
            ON company_relation_documents(parent_company_id, child_company_id, document_kind);
            """
        )
        self.conn.commit()

    def fetchall(self, query: str, params: tuple = ()) -> list[sqlite3.Row]:
        return list(self.conn.execute(query, params))

    def fetchone(self, query: str, params: tuple = ()) -> sqlite3.Row | None:
        return self.conn.execute(query, params).fetchone()

    def execute(self, query: str, params: tuple = ()) -> sqlite3.Cursor:
        cur = self.conn.execute(query, params)
        self.conn.commit()
        return cur

    def create_project(self, data: dict[str, str]) -> int:
        ts = now_text()
        cur = self.execute(
            """
            INSERT INTO projects (
                name, construction_no, client_name, start_date, end_date,
                site_agent, managing_engineer, chief_engineer, base_folder,
                created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data["name"],
                data["construction_no"],
                data["client_name"],
                data["start_date"],
                data["end_date"],
                data["site_agent"],
                data["managing_engineer"],
                data["chief_engineer"],
                data["base_folder"],
                ts,
                ts,
            ),
        )
        return int(cur.lastrowid)

    def list_projects(self) -> list[sqlite3.Row]:
        return self.fetchall(
            """
            SELECT *
            FROM projects
            ORDER BY updated_at DESC, id DESC
            """
        )

    def find_project_by_construction_no(self, construction_no: str) -> sqlite3.Row | None:
        return self.fetchone(
            """
            SELECT *
            FROM projects
            WHERE construction_no = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (construction_no,),
        )

    def update_project(self, project_id: int, data: dict[str, str]) -> None:
        self.execute(
            """
            UPDATE projects
            SET name = ?,
                construction_no = ?,
                client_name = ?,
                start_date = ?,
                end_date = ?,
                site_agent = ?,
                managing_engineer = ?,
                chief_engineer = ?,
                base_folder = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                data["name"],
                data["construction_no"],
                data["client_name"],
                data["start_date"],
                data["end_date"],
                data["site_agent"],
                data["managing_engineer"],
                data["chief_engineer"],
                data["base_folder"],
                now_text(),
                project_id,
            ),
        )

    def delete_project(self, project_id: int) -> None:
        self.execute("DELETE FROM projects WHERE id = ?", (project_id,))

    def reset_all_data(self) -> None:
        self.conn.execute("PRAGMA foreign_keys = OFF")
        try:
            self.conn.execute("DELETE FROM relation_document_attachments")
            self.conn.execute("DELETE FROM company_relation_documents")
            self.conn.execute("DELETE FROM attachments")
            self.conn.execute("DELETE FROM required_documents")
            self.conn.execute("DELETE FROM companies")
            self.conn.execute("DELETE FROM project_imports")
            self.conn.execute("DELETE FROM projects")
            self.conn.commit()
        finally:
            self.conn.execute("PRAGMA foreign_keys = ON")

    def find_project_by_name(self, name: str) -> sqlite3.Row | None:
        return self.fetchone(
            """
            SELECT *
            FROM projects
            WHERE name = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (name,),
        )

    def upsert_project_import(self, project_id: int, import_kind: str, source_path: str, source_sheet: str) -> None:
        row = self.fetchone(
            """
            SELECT id
            FROM project_imports
            WHERE project_id = ? AND import_kind = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (project_id, import_kind),
        )
        ts = now_text()
        if row is None:
            self.execute(
                """
                INSERT INTO project_imports (
                    project_id, import_kind, source_path, source_sheet, imported_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?)
                """,
                (project_id, import_kind, source_path, source_sheet, ts, ts),
            )
            return
        self.execute(
            """
            UPDATE project_imports
            SET source_path = ?, source_sheet = ?, updated_at = ?
            WHERE id = ?
            """,
            (source_path, source_sheet, ts, row["id"]),
        )

    def get_latest_project_import(self, project_id: int, import_kind: str) -> sqlite3.Row | None:
        return self.fetchone(
            """
            SELECT *
            FROM project_imports
            WHERE project_id = ? AND import_kind = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (project_id, import_kind),
        )

    def ensure_relation_document_for_child(self, child_company_id: int) -> int | None:
        child = self.get_company(child_company_id)
        if child is None or child["parent_company_id"] is None:
            return None
        parent = self.get_company(child["parent_company_id"])
        if parent is None:
            return None
        kind = relation_document_kind(int(parent["level"]), int(child["level"]))
        row = self.fetchone(
            """
            SELECT id
            FROM company_relation_documents
            WHERE parent_company_id = ?
              AND child_company_id = ?
              AND document_kind = ?
            ORDER BY id DESC
            LIMIT 1
            """,
            (parent["id"], child["id"], kind),
        )
        if row is not None:
            return int(row["id"])
        ts = now_text()
        cur = self.execute(
            """
            INSERT INTO company_relation_documents (
                project_id, parent_company_id, child_company_id, document_kind,
                status, source_path, source_sheet, reflected_path, note, created_at, updated_at
            ) VALUES (?, ?, ?, ?, '未確認', '', '', '', '', ?, ?)
            """,
            (child["project_id"], parent["id"], child["id"], kind, ts, ts),
        )
        return int(cur.lastrowid)

    def upsert_relation_document_import(self, child_company_id: int, source_path: str, source_sheet: str) -> int | None:
        relation_document_id = self.ensure_relation_document_for_child(child_company_id)
        if relation_document_id is None:
            return None
        self.update_relation_document_import(relation_document_id, source_path, source_sheet)
        return relation_document_id

    def update_relation_document_import(self, relation_document_id: int, source_path: str, source_sheet: str) -> None:
        self.execute(
            """
            UPDATE company_relation_documents
            SET status = '添付済み',
                source_path = ?,
                source_sheet = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (source_path, source_sheet, now_text(), relation_document_id),
        )

    def update_relation_document_reflected_path(self, relation_document_id: int, reflected_path: str) -> None:
        self.execute(
            """
            UPDATE company_relation_documents
            SET reflected_path = ?, updated_at = ?
            WHERE id = ?
            """,
            (reflected_path, now_text(), relation_document_id),
        )

    def update_relation_document_status(self, relation_document_id: int, status: str) -> None:
        self.execute(
            """
            UPDATE company_relation_documents
            SET status = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, now_text(), relation_document_id),
        )

    def update_relation_document_note(self, relation_document_id: int, status: str, note: str) -> None:
        self.execute(
            """
            UPDATE company_relation_documents
            SET status = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, note, now_text(), relation_document_id),
        )

    def clear_relation_document_registration(self, relation_document_id: int) -> None:
        self.execute("DELETE FROM relation_document_attachments WHERE relation_document_id = ?", (relation_document_id,))
        self.execute(
            """
            UPDATE company_relation_documents
            SET status = '未確認',
                source_path = '',
                source_sheet = '',
                reflected_path = '',
                note = '',
                updated_at = ?
            WHERE id = ?
            """,
            (now_text(), relation_document_id),
        )

    def list_relation_documents(self, project_id: int) -> list[sqlite3.Row]:
        return self.fetchall(
            """
            SELECT rd.*,
                   parent.name AS parent_company_name,
                   child.name AS child_company_name,
                   COUNT(a.id) AS attachment_count
            FROM company_relation_documents rd
            INNER JOIN companies parent ON parent.id = rd.parent_company_id
            INNER JOIN companies child ON child.id = rd.child_company_id
            LEFT JOIN relation_document_attachments a ON a.relation_document_id = rd.id
            WHERE rd.project_id = ?
            GROUP BY rd.id
            ORDER BY parent.level ASC, parent.name ASC, child.name ASC, rd.id ASC
            """,
            (project_id,),
        )

    def get_relation_document(self, relation_document_id: int) -> sqlite3.Row | None:
        return self.fetchone(
            """
            SELECT rd.*,
                   parent.name AS parent_company_name,
                   parent.level AS parent_level,
                   child.name AS child_company_name,
                   child.level AS child_level
            FROM company_relation_documents rd
            INNER JOIN companies parent ON parent.id = rd.parent_company_id
            INNER JOIN companies child ON child.id = rd.child_company_id
            WHERE rd.id = ?
            """,
            (relation_document_id,),
        )

    def get_relation_document_for_child(self, child_company_id: int) -> sqlite3.Row | None:
        relation_document_id = self.ensure_relation_document_for_child(child_company_id)
        if relation_document_id is None:
            return None
        return self.get_relation_document(relation_document_id)

    def create_company(self, data: dict[str, str | int | None]) -> int:
        ts = now_text()
        cur = self.execute(
            """
            INSERT INTO companies (
                project_id, parent_company_id, level, name, kana, representative,
                address, phone, license_no, license_expiry, work_type, contract_date,
                planned_start_date, planned_end_date, chief_engineer_name,
                chief_engineer_license, safety_manager, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                data["project_id"],
                data["parent_company_id"],
                data["level"],
                data["name"],
                data["kana"],
                data["representative"],
                data["address"],
                data["phone"],
                data["license_no"],
                data["license_expiry"],
                data["work_type"],
                data["contract_date"],
                data["planned_start_date"],
                data["planned_end_date"],
                data["chief_engineer_name"],
                data["chief_engineer_license"],
                data["safety_manager"],
                ts,
                ts,
            ),
        )
        company_id = int(cur.lastrowid)
        self.generate_required_documents(company_id)
        return company_id

    def update_company(self, company_id: int, data: dict[str, str]) -> None:
        self.execute(
            """
            UPDATE companies
            SET name = ?,
                kana = ?,
                representative = ?,
                address = ?,
                phone = ?,
                license_no = ?,
                license_expiry = ?,
                work_type = ?,
                contract_date = ?,
                planned_start_date = ?,
                planned_end_date = ?,
                chief_engineer_name = ?,
                chief_engineer_license = ?,
                safety_manager = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (
                data["name"],
                data["kana"],
                data["representative"],
                data["address"],
                data["phone"],
                data["license_no"],
                data["license_expiry"],
                data["work_type"],
                data["contract_date"],
                data["planned_start_date"],
                data["planned_end_date"],
                data["chief_engineer_name"],
                data["chief_engineer_license"],
                data["safety_manager"],
                now_text(),
                company_id,
            ),
        )

    def generate_required_documents(self, company_id: int) -> None:
        ts = now_text()
        existing = {
            row["document_name"]
            for row in self.fetchall(
                "SELECT document_name FROM required_documents WHERE company_id = ?",
                (company_id,),
            )
        }
        for name in DOCUMENT_TEMPLATES:
            if name in existing:
                continue
            self.execute(
                """
                INSERT INTO required_documents (
                    company_id, document_name, required, status, expiry_date, note, created_at, updated_at
                ) VALUES (?, ?, 1, '未確認', '', '', ?, ?)
                """,
                (company_id, name, ts, ts),
            )

    def list_companies(self, project_id: int) -> list[sqlite3.Row]:
        return self.fetchall(
            """
            SELECT *
            FROM companies
            WHERE project_id = ?
            ORDER BY level ASC, id ASC
            """,
            (project_id,),
        )

    def get_company(self, company_id: int) -> sqlite3.Row | None:
        return self.fetchone("SELECT * FROM companies WHERE id = ?", (company_id,))

    def list_required_documents(self, company_id: int) -> list[sqlite3.Row]:
        return self.fetchall(
            """
            SELECT rd.*,
                   COUNT(a.id) AS attachment_count
            FROM required_documents rd
            LEFT JOIN attachments a ON a.required_document_id = rd.id
            WHERE rd.company_id = ?
            GROUP BY rd.id
            ORDER BY rd.id ASC
            """,
            (company_id,),
        )

    def update_required_document(self, document_id: int, status: str, expiry_date: str, note: str) -> None:
        self.execute(
            """
            UPDATE required_documents
            SET status = ?, expiry_date = ?, note = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, expiry_date, note, now_text(), document_id),
        )

    def list_shortages(self, project_id: int) -> list[sqlite3.Row]:
        return self.fetchall(
            """
            SELECT company_name, document_name, status
            FROM (
                SELECT c.level AS sort_level,
                       c.name AS sort_name,
                       c.name AS company_name,
                       rd.document_name,
                       rd.status
                FROM required_documents rd
                INNER JOIN companies c ON c.id = rd.company_id
                WHERE c.project_id = ?
                  AND rd.status IN ('不足', '未確認', '期限切れ')
                UNION ALL
                SELECT child.level AS sort_level,
                       parent.name || ' -> ' || child.name AS sort_name,
                       parent.name || ' -> ' || child.name AS company_name,
                       rel.document_kind AS document_name,
                       rel.status
                FROM company_relation_documents rel
                INNER JOIN companies parent ON parent.id = rel.parent_company_id
                INNER JOIN companies child ON child.id = rel.child_company_id
                WHERE rel.project_id = ?
                  AND rel.status IN ('不足', '未確認', '期限切れ')
            )
            ORDER BY sort_level ASC, sort_name ASC, document_name ASC
            """,
            (project_id, project_id),
        )

    def get_project(self, project_id: int) -> sqlite3.Row | None:
        return self.fetchone("SELECT * FROM projects WHERE id = ?", (project_id,))

    def add_attachment(self, document_id: int, original_path: Path, stored_path: Path) -> int:
        cur = self.execute(
            """
            INSERT INTO attachments (
                required_document_id, original_path, stored_path, original_filename,
                stored_filename, file_type, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                document_id,
                str(original_path),
                str(stored_path),
                original_path.name,
                stored_path.name,
                original_path.suffix.lower().lstrip("."),
                now_text(),
            ),
        )
        return int(cur.lastrowid)

    def list_attachments(self, document_id: int) -> list[sqlite3.Row]:
        return self.fetchall(
            """
            SELECT *
            FROM attachments
            WHERE required_document_id = ?
            ORDER BY id DESC
            """,
            (document_id,),
        )

    def get_attachment(self, attachment_id: int) -> sqlite3.Row | None:
        return self.fetchone("SELECT * FROM attachments WHERE id = ?", (attachment_id,))

    def add_relation_attachment(self, relation_document_id: int, original_path: Path, stored_path: Path) -> int:
        cur = self.execute(
            """
            INSERT INTO relation_document_attachments (
                relation_document_id, original_path, stored_path, original_filename,
                stored_filename, file_type, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                relation_document_id,
                str(original_path),
                str(stored_path),
                original_path.name,
                stored_path.name,
                original_path.suffix.lower().lstrip("."),
                now_text(),
            ),
        )
        return int(cur.lastrowid)

    def list_relation_attachments(self, relation_document_id: int) -> list[sqlite3.Row]:
        return self.fetchall(
            """
            SELECT *
            FROM relation_document_attachments
            WHERE relation_document_id = ?
            ORDER BY id DESC
            """,
            (relation_document_id,),
        )

    def get_relation_attachment(self, attachment_id: int) -> sqlite3.Row | None:
        return self.fetchone("SELECT * FROM relation_document_attachments WHERE id = ?", (attachment_id,))

    def delete_relation_attachment(self, attachment_id: int) -> None:
        self.execute("DELETE FROM relation_document_attachments WHERE id = ?", (attachment_id,))

    def next_versioned_filename(self, base_dir: Path, company_name: str, document_name: str, suffix: str) -> str:
        safe_company = sanitize_filename(company_name)
        safe_document = sanitize_filename(document_name)
        version = 1
        while True:
            candidate = f"{safe_company}_{safe_document}_v{version}{suffix}"
            if not (base_dir / candidate).exists():
                return candidate
            version += 1


@dataclass
class ProjectSelection:
    project_id: int | None = None
    company_id: int | None = None
    document_id: int | None = None
    attachment_id: int | None = None
    relation_document_id: int | None = None
    relation_attachment_id: int | None = None


class ProjectDialog:
    def __init__(self, master: tk.Misc, title: str = "工事登録", initial_data: dict[str, str] | None = None) -> None:
        self.top = Toplevel(master)
        self.top.title(title)
        self.top.transient(master)
        self.top.grab_set()
        self.result: dict[str, str] | None = None

        fields = [
            ("工事名", "name"),
            ("工事番号", "construction_no"),
            ("発注者", "client_name"),
            ("工期開始日", "start_date"),
            ("工期終了日", "end_date"),
            ("現場代理人", "site_agent"),
            ("監理技術者", "managing_engineer"),
            ("主任技術者", "chief_engineer"),
            ("保存フォルダ", "base_folder"),
        ]
        self.vars: dict[str, StringVar] = {}
        for index, (label, key) in enumerate(fields):
            ttk.Label(self.top, text=label).grid(row=index, column=0, padx=8, pady=4, sticky="w")
            var = StringVar(value=(initial_data or {}).get(key, ""))
            self.vars[key] = var
            entry = ttk.Entry(self.top, textvariable=var, width=42)
            entry.grid(row=index, column=1, padx=8, pady=4, sticky="ew")
            if key == "base_folder":
                ttk.Button(self.top, text="参照", command=self.pick_folder).grid(row=index, column=2, padx=8, pady=4)

        button_row = len(fields)
        ttk.Button(self.top, text="保存", command=self.submit).grid(row=button_row, column=1, padx=8, pady=10, sticky="e")
        ttk.Button(self.top, text="キャンセル", command=self.top.destroy).grid(row=button_row, column=2, padx=8, pady=10, sticky="w")
        self.top.columnconfigure(1, weight=1)

    def pick_folder(self) -> None:
        folder = filedialog.askdirectory(title="保存フォルダを選択")
        if folder:
            self.vars["base_folder"].set(folder)

    def submit(self) -> None:
        required = ["name", "construction_no", "client_name"]
        for key in required:
            if not self.vars[key].get().strip():
                messagebox.showerror("入力エラー", "工事名・工事番号・発注者は必須です。", parent=self.top)
                return
        self.result = {key: var.get().strip() for key, var in self.vars.items()}
        self.top.destroy()


class CompanyDialog:
    def __init__(
        self,
        master: tk.Misc,
        level: int,
        parent_name: str | None,
        title: str = "業者登録",
        initial_data: dict[str, str] | None = None,
    ) -> None:
        self.top = Toplevel(master)
        self.top.title(title)
        self.top.transient(master)
        self.top.grab_set()
        self.result: dict[str, str] | None = None

        title = LEVEL_LABELS.get(level, "下請")
        header = title if parent_name is None else f"{title} / 親: {parent_name}"
        ttk.Label(self.top, text=header).grid(row=0, column=0, columnspan=2, padx=8, pady=(8, 12), sticky="w")

        fields = [
            ("会社名", "name"),
            ("会社名カナ", "kana"),
            ("代表者名", "representative"),
            ("所在地", "address"),
            ("電話番号", "phone"),
            ("建設業許可番号", "license_no"),
            ("許可有効期限", "license_expiry"),
            ("担当工種", "work_type"),
            ("契約日", "contract_date"),
            ("施工開始予定日", "planned_start_date"),
            ("施工終了予定日", "planned_end_date"),
            ("主任技術者名", "chief_engineer_name"),
            ("主任技術者資格", "chief_engineer_license"),
            ("安全衛生責任者", "safety_manager"),
        ]
        self.vars: dict[str, StringVar] = {}
        for index, (label, key) in enumerate(fields, start=1):
            ttk.Label(self.top, text=label).grid(row=index, column=0, padx=8, pady=4, sticky="w")
            var = StringVar(value=(initial_data or {}).get(key, ""))
            self.vars[key] = var
            ttk.Entry(self.top, textvariable=var, width=40).grid(row=index, column=1, padx=8, pady=4, sticky="ew")

        button_row = len(fields) + 1
        ttk.Button(self.top, text="保存", command=self.submit).grid(row=button_row, column=1, padx=8, pady=10, sticky="e")
        ttk.Button(self.top, text="キャンセル", command=self.top.destroy).grid(row=button_row, column=0, padx=8, pady=10, sticky="w")
        self.top.columnconfigure(1, weight=1)

    def submit(self) -> None:
        if not self.vars["name"].get().strip():
            messagebox.showerror("入力エラー", "会社名は必須です。", parent=self.top)
            return
        self.result = {key: var.get().strip() for key, var in self.vars.items()}
        self.top.destroy()


class DocumentStatusDialog:
    def __init__(self, master: tk.Misc, document: sqlite3.Row) -> None:
        self.top = Toplevel(master)
        self.top.title("書類状態更新")
        self.top.transient(master)
        self.top.grab_set()
        self.result: dict[str, str] | None = None

        ttk.Label(self.top, text=document["document_name"]).grid(row=0, column=0, columnspan=2, padx=8, pady=(8, 12), sticky="w")

        self.status_var = StringVar(value=document["status"])
        self.expiry_var = StringVar(value=document["expiry_date"] or "")
        self.note_var = StringVar(value=document["note"] or "")

        ttk.Label(self.top, text="状態").grid(row=1, column=0, padx=8, pady=4, sticky="w")
        ttk.Combobox(self.top, textvariable=self.status_var, values=DOCUMENT_STATUSES, state="readonly").grid(row=1, column=1, padx=8, pady=4, sticky="ew")

        ttk.Label(self.top, text="有効期限").grid(row=2, column=0, padx=8, pady=4, sticky="w")
        ttk.Entry(self.top, textvariable=self.expiry_var).grid(row=2, column=1, padx=8, pady=4, sticky="ew")

        ttk.Label(self.top, text="備考").grid(row=3, column=0, padx=8, pady=4, sticky="nw")
        self.note_text = tk.Text(self.top, width=36, height=6)
        self.note_text.grid(row=3, column=1, padx=8, pady=4, sticky="ew")
        self.note_text.insert("1.0", document["note"] or "")

        ttk.Button(self.top, text="保存", command=self.submit).grid(row=4, column=1, padx=8, pady=10, sticky="e")
        ttk.Button(self.top, text="キャンセル", command=self.top.destroy).grid(row=4, column=0, padx=8, pady=10, sticky="w")
        self.top.columnconfigure(1, weight=1)

    def submit(self) -> None:
        self.result = {
            "status": self.status_var.get(),
            "expiry_date": self.expiry_var.get().strip(),
            "note": self.note_text.get("1.0", END).strip(),
        }
        self.top.destroy()


class RelationDocumentDialog:
    def __init__(self, master: tk.Misc, relation_document: sqlite3.Row) -> None:
        self.top = Toplevel(master)
        self.top.title("帳票情報編集")
        self.top.transient(master)
        self.top.grab_set()
        self.result: dict[str, str] | None = None

        title = f"{relation_document['parent_company_name']} -> {relation_document['child_company_name']}"
        ttk.Label(self.top, text=title).grid(row=0, column=0, columnspan=2, padx=8, pady=(8, 4), sticky="w")
        ttk.Label(self.top, text=relation_document["document_kind"]).grid(row=1, column=0, columnspan=2, padx=8, pady=(0, 12), sticky="w")

        self.status_var = StringVar(value=relation_document["status"])

        ttk.Label(self.top, text="状態").grid(row=2, column=0, padx=8, pady=4, sticky="w")
        ttk.Combobox(self.top, textvariable=self.status_var, values=DOCUMENT_STATUSES, state="readonly").grid(row=2, column=1, padx=8, pady=4, sticky="ew")

        ttk.Label(self.top, text="備考").grid(row=3, column=0, padx=8, pady=4, sticky="nw")
        self.note_text = tk.Text(self.top, width=42, height=6)
        self.note_text.grid(row=3, column=1, padx=8, pady=4, sticky="ew")
        self.note_text.insert("1.0", relation_document["note"] or "")

        ttk.Button(self.top, text="保存", command=self.submit).grid(row=4, column=1, padx=8, pady=10, sticky="e")
        ttk.Button(self.top, text="キャンセル", command=self.top.destroy).grid(row=4, column=0, padx=8, pady=10, sticky="w")
        self.top.columnconfigure(1, weight=1)

    def submit(self) -> None:
        self.result = {
            "status": self.status_var.get(),
            "note": self.note_text.get("1.0", END).strip(),
        }
        self.top.destroy()


class CsvImportConfirmDialog:
    def __init__(self, master: tk.Misc, rows: list[dict[str, str]]) -> None:
        self.top = Toplevel(master)
        self.top.title("CSV取込確認")
        self.top.transient(master)
        self.top.grab_set()
        self.approved = False

        ttk.Label(
            self.top,
            text="工事番号で照合した取込候補です。内容を確認して取り込みを実行してください。",
        ).pack(anchor="w", padx=10, pady=(10, 8))

        columns = ("action", "construction_no", "name", "client_name", "start_date", "end_date")
        tree = ttk.Treeview(self.top, columns=columns, show="headings", height=12)
        tree.heading("action", text="取込区分")
        tree.heading("construction_no", text="工事番号")
        tree.heading("name", text="工事名")
        tree.heading("client_name", text="発注者")
        tree.heading("start_date", text="開始日")
        tree.heading("end_date", text="終了日")
        tree.column("action", width=80, anchor="center")
        tree.column("construction_no", width=120, anchor="center")
        tree.column("name", width=240)
        tree.column("client_name", width=180)
        tree.column("start_date", width=100, anchor="center")
        tree.column("end_date", width=100, anchor="center")
        tree.pack(fill=BOTH, expand=True, padx=10, pady=(0, 10))

        for row in rows:
            tree.insert(
                "",
                END,
                values=(
                    row["action"],
                    row["construction_no"],
                    row["name"],
                    row["client_name"],
                    row["start_date"],
                    row["end_date"],
                ),
            )

        buttons = ttk.Frame(self.top)
        buttons.pack(fill=X, padx=10, pady=(0, 10))
        ttk.Button(buttons, text="取り込む", command=self.submit).pack(side=RIGHT)
        ttk.Button(buttons, text="キャンセル", command=self.top.destroy).pack(side=RIGHT, padx=(0, 8))

    def submit(self) -> None:
        self.approved = True
        self.top.destroy()


class CivilHubApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.db = Database(DB_PATH)
        self.selection = ProjectSelection()
        self.project_map: dict[str, int] = {}
        self.company_tree_map: dict[str, int] = {}
        self.document_tree_map: dict[str, int] = {}
        self.attachment_tree_map: dict[str, int] = {}
        self.relation_document_tree_map: dict[str, int] = {}
        self.relation_attachment_tree_map: dict[str, int] = {}
        self.last_ledger_import_items: list[dict[str, str]] = []

        self.root.title(f"{APP_TITLE} - {APP_SUBTITLE}")
        self.root.geometry("1500x900")

        self.style = ttk.Style(self.root)
        if "clam" in self.style.theme_names():
            self.style.theme_use("clam")

        self.build_ui()
        self.refresh_projects()

    def build_ui(self) -> None:
        header = ttk.Frame(self.root, padding=10)
        header.pack(fill=X)
        ttk.Label(header, text=APP_TITLE, font=("Yu Gothic UI", 20, "bold")).pack(side=LEFT)
        ttk.Label(header, text=APP_SUBTITLE, font=("Yu Gothic UI", 12)).pack(side=LEFT, padx=(12, 0))
        ttk.Button(header, text="不足一覧", command=self.show_shortages).pack(side=RIGHT)

        body = ttk.Panedwindow(self.root, orient=tk.HORIZONTAL)
        body.pack(fill=BOTH, expand=True)

        left = ttk.Frame(body, padding=10)
        center = ttk.Frame(body, padding=10)
        right = ttk.Frame(body, padding=10)
        body.add(left, weight=2)
        body.add(center, weight=3)
        body.add(right, weight=3)

        self.build_project_panel(left)
        self.build_company_panel(center)
        self.build_detail_panel(right)

        bottom = ttk.Notebook(self.root)
        bottom.pack(fill=BOTH, expand=True, padx=10, pady=(0, 10))
        docs_tab = ttk.Frame(bottom, padding=10)
        attach_tab = ttk.Frame(bottom, padding=10)
        relation_tab = ttk.Frame(bottom, padding=10)
        bottom.add(docs_tab, text="添付書類チェック")
        bottom.add(attach_tab, text="ファイル添付")
        bottom.add(relation_tab, text="契約関係帳票")
        self.build_document_panel(docs_tab)
        self.build_attachment_panel(attach_tab)
        self.build_relation_panel(relation_tab)

        footer = ttk.Frame(self.root, padding=(10, 4))
        footer.pack(fill=X)
        self.status_var = StringVar(value=f"DB: {DB_PATH.name}")
        ttk.Label(footer, textvariable=self.status_var).pack(side=LEFT)
        ttk.Button(footer, text="開発用初期化", command=self.reset_debug_data).pack(side=RIGHT)

    def build_project_panel(self, parent: ttk.Frame) -> None:
        top = ttk.Frame(parent)
        top.pack(fill=X)
        ttk.Label(top, text="工事一覧", font=("Yu Gothic UI", 12, "bold")).pack(side=LEFT)

        ttk.Label(parent, text="工事を選択して、施工体制と添付書類を管理します。").pack(anchor="w", pady=(8, 0))

        primary_buttons = ttk.Frame(parent)
        primary_buttons.pack(fill=X, pady=(8, 0))
        ttk.Button(primary_buttons, text="+ 工事登録", command=self.add_project).pack(side=LEFT)
        ttk.Button(primary_buttons, text="工事編集", command=self.edit_selected_project).pack(side=LEFT, padx=(6, 0))

        ledger_buttons = ttk.Frame(parent)
        ledger_buttons.pack(fill=X, pady=(6, 0))
        ttk.Button(ledger_buttons, text="工事情報Excel取込", command=self.import_ledger_excel).pack(side=LEFT)
        ttk.Button(ledger_buttons, text="工事情報反映コピー", command=self.reflect_project_to_ledger).pack(side=LEFT, padx=(6, 0))

        support_buttons = ttk.Frame(parent)
        support_buttons.pack(fill=X, pady=(6, 0))
        ttk.Button(support_buttons, text="工事再読込", command=self.refresh_projects).pack(side=LEFT)
        ttk.Button(support_buttons, text="CSV取込", command=self.import_projects_csv).pack(side=LEFT, padx=(6, 0))
        ttk.Button(support_buttons, text="工事削除", command=self.delete_selected_project).pack(side=LEFT, padx=(6, 0))

        self.project_list = tk.Listbox(parent, height=18)
        self.project_list.pack(fill=BOTH, expand=True, pady=(8, 8))
        self.project_list.bind("<<ListboxSelect>>", self.on_project_selected)

    def build_company_panel(self, parent: ttk.Frame) -> None:
        top = ttk.Frame(parent)
        top.pack(fill=X)
        ttk.Label(top, text="施工体制ツリー", font=("Yu Gothic UI", 12, "bold")).pack(side=LEFT)
        ttk.Button(top, text="+ 元請", command=self.add_root_company).pack(side=RIGHT)
        ttk.Button(top, text="+ 子業者", command=self.add_child_company).pack(side=RIGHT, padx=(0, 6))
        ttk.Button(top, text="会社情報編集", command=self.edit_selected_company).pack(side=RIGHT, padx=(0, 6))
        ttk.Label(parent, text="工事を選択し、元請から順に施工体制を登録します。").pack(anchor="w", pady=(8, 0))

        columns = ("level", "work_type")
        self.company_tree = ttk.Treeview(parent, columns=columns, show="tree headings", height=22)
        self.company_tree.heading("#0", text="会社名")
        self.company_tree.heading("level", text="階層")
        self.company_tree.heading("work_type", text="担当工種")
        self.company_tree.column("#0", width=260)
        self.company_tree.column("level", width=100, anchor="center")
        self.company_tree.column("work_type", width=140)
        self.company_tree.pack(fill=BOTH, expand=True, pady=(6, 0))
        self.company_tree.bind("<<TreeviewSelect>>", self.on_company_selected)

    def build_detail_panel(self, parent: ttk.Frame) -> None:
        ttk.Label(parent, text="選択業者の詳細", font=("Yu Gothic UI", 12, "bold")).pack(anchor="w")
        ttk.Label(parent, text="施工体制ツリーで業者を選択すると詳細を確認できます。").pack(anchor="w", pady=(8, 0))
        self.detail_text = tk.Text(parent, width=42, height=26)
        self.detail_text.pack(fill=BOTH, expand=True, pady=(6, 8))
        self.detail_text.configure(state="disabled")

        button_row = ttk.Frame(parent)
        button_row.pack(fill=X)
        ttk.Button(button_row, text="書類状態更新", command=self.edit_selected_document).pack(side=LEFT)
        ttk.Button(button_row, text="リネーム候補", command=self.show_rename_candidate).pack(side=LEFT, padx=(8, 0))

    def build_document_panel(self, parent: ttk.Frame) -> None:
        top = ttk.Frame(parent)
        top.pack(fill=X)
        ttk.Label(top, text="必要添付書類", font=("Yu Gothic UI", 11, "bold")).pack(side=LEFT)
        ttk.Button(top, text="状態更新", command=self.edit_selected_document).pack(side=RIGHT)
        ttk.Label(parent, text="不足・未確認・期限切れを優先して確認してください。行のダブルクリックでも状態更新できます。").pack(anchor="w", pady=(8, 0))

        columns = ("document_name", "required", "status", "expiry_date", "attachment_count", "note")
        self.document_tree = ttk.Treeview(parent, columns=columns, show="headings", height=12)
        self.document_tree.heading("document_name", text="書類名")
        self.document_tree.heading("required", text="必須")
        self.document_tree.heading("status", text="状態")
        self.document_tree.heading("expiry_date", text="有効期限")
        self.document_tree.heading("attachment_count", text="添付数")
        self.document_tree.heading("note", text="備考")
        self.document_tree.column("document_name", width=240)
        self.document_tree.column("required", width=60, anchor="center")
        self.document_tree.column("status", width=100, anchor="center")
        self.document_tree.column("expiry_date", width=120, anchor="center")
        self.document_tree.column("attachment_count", width=100, anchor="center")
        self.document_tree.column("note", width=380)
        self.document_tree.pack(fill=BOTH, expand=True, pady=(6, 0))
        self.document_tree.bind("<<TreeviewSelect>>", self.on_document_selected)
        self.document_tree.bind("<Double-1>", lambda _event: self.edit_selected_document())

        for status, color in STATUS_COLORS.items():
            self.document_tree.tag_configure(status, background=color, foreground="white")

    def build_attachment_panel(self, parent: ttk.Frame) -> None:
        top = ttk.Frame(parent)
        top.pack(fill=X)
        ttk.Label(top, text="添付ファイル", font=("Yu Gothic UI", 11, "bold")).pack(side=LEFT)
        ttk.Button(top, text="+ 添付", command=self.attach_file).pack(side=RIGHT)
        ttk.Button(top, text="保存場所を開く", command=self.open_attachment_folder).pack(side=RIGHT, padx=(0, 6))
        ttk.Button(top, text="ファイルを開く", command=self.open_attachment).pack(side=RIGHT, padx=(0, 6))

        columns = ("original_filename", "stored_filename", "file_type", "created_at")
        self.attachment_tree = ttk.Treeview(parent, columns=columns, show="headings", height=12)
        self.attachment_tree.heading("original_filename", text="元ファイル名")
        self.attachment_tree.heading("stored_filename", text="保存ファイル名")
        self.attachment_tree.heading("file_type", text="種類")
        self.attachment_tree.heading("created_at", text="登録日時")
        self.attachment_tree.column("original_filename", width=280)
        self.attachment_tree.column("stored_filename", width=280)
        self.attachment_tree.column("file_type", width=80, anchor="center")
        self.attachment_tree.column("created_at", width=150, anchor="center")
        self.attachment_tree.pack(fill=BOTH, expand=True, pady=(8, 0))
        self.attachment_tree.bind("<<TreeviewSelect>>", self.on_attachment_selected)

    def build_relation_panel(self, parent: ttk.Frame) -> None:
        top = ttk.Frame(parent)
        top.pack(fill=X)
        ttk.Label(top, text="契約関係帳票", font=("Yu Gothic UI", 11, "bold")).pack(side=LEFT)
        ttk.Button(top, text="帳票取込", command=self.import_relation_document_excel).pack(side=RIGHT)
        ttk.Button(top, text="+ 帳票添付", command=self.attach_relation_document_file).pack(side=RIGHT, padx=(0, 6))
        ttk.Label(
            parent,
            text="施工体制ツリーで子業者を選択すると、親 -> 子 の帳票を操作できます。複数シートExcelは先頭シートを対象にします。",
        ).pack(anchor="w", pady=(8, 0))

        self.relation_summary_var = StringVar(value="子業者を選択すると、この親子関係に必要な帳票を表示します。")
        ttk.Label(parent, textvariable=self.relation_summary_var).pack(anchor="w", pady=(6, 0))

        detail_buttons = ttk.Frame(parent)
        detail_buttons.pack(fill=X, pady=(6, 0))
        ttk.Button(detail_buttons, text="帳票情報編集", command=self.edit_relation_document_info).pack(side=LEFT)
        ttk.Button(detail_buttons, text="取込元差し替え", command=self.replace_relation_document_file).pack(side=LEFT, padx=(6, 0))
        ttk.Button(detail_buttons, text="登録削除", command=self.clear_relation_document_registration).pack(side=LEFT, padx=(6, 0))
        ttk.Button(detail_buttons, text="反映コピー作成", command=self.reflect_relation_document).pack(side=LEFT, padx=(6, 0))
        ttk.Button(detail_buttons, text="保存場所を開く", command=self.open_relation_attachment_folder).pack(side=LEFT, padx=(6, 0))
        ttk.Button(detail_buttons, text="ファイルを開く", command=self.open_relation_attachment).pack(side=LEFT, padx=(6, 0))
        ttk.Button(detail_buttons, text="添付登録削除", command=self.delete_relation_attachment_registration).pack(side=LEFT, padx=(6, 0))

        columns = ("parent_child", "document_kind", "status", "source_file", "updated_at", "attachment_count")
        self.relation_document_tree = ttk.Treeview(parent, columns=columns, show="headings", height=7)
        self.relation_document_tree.heading("parent_child", text="親 -> 子")
        self.relation_document_tree.heading("document_kind", text="帳票種別")
        self.relation_document_tree.heading("status", text="状態")
        self.relation_document_tree.heading("source_file", text="取込元Excel")
        self.relation_document_tree.heading("updated_at", text="最終更新日時")
        self.relation_document_tree.heading("attachment_count", text="添付数")
        self.relation_document_tree.column("parent_child", width=260)
        self.relation_document_tree.column("document_kind", width=150, anchor="center")
        self.relation_document_tree.column("status", width=90, anchor="center")
        self.relation_document_tree.column("source_file", width=220)
        self.relation_document_tree.column("updated_at", width=150, anchor="center")
        self.relation_document_tree.column("attachment_count", width=80, anchor="center")
        self.relation_document_tree.pack(fill=BOTH, expand=True, pady=(6, 8))
        self.relation_document_tree.bind("<<TreeviewSelect>>", self.on_relation_document_selected)
        for status, color in STATUS_COLORS.items():
            self.relation_document_tree.tag_configure(status, background=color, foreground="white")

        ttk.Label(parent, text="契約関係帳票の添付ファイル", font=("Yu Gothic UI", 11, "bold")).pack(anchor="w")
        attach_columns = ("original_filename", "stored_filename", "file_type", "created_at")
        self.relation_attachment_tree = ttk.Treeview(parent, columns=attach_columns, show="headings", height=5)
        self.relation_attachment_tree.heading("original_filename", text="元ファイル名")
        self.relation_attachment_tree.heading("stored_filename", text="保存ファイル名")
        self.relation_attachment_tree.heading("file_type", text="種類")
        self.relation_attachment_tree.heading("created_at", text="登録日時")
        self.relation_attachment_tree.column("original_filename", width=280)
        self.relation_attachment_tree.column("stored_filename", width=280)
        self.relation_attachment_tree.column("file_type", width=80, anchor="center")
        self.relation_attachment_tree.column("created_at", width=150, anchor="center")
        self.relation_attachment_tree.pack(fill=BOTH, expand=True, pady=(6, 0))
        self.relation_attachment_tree.bind("<<TreeviewSelect>>", self.on_relation_attachment_selected)

    def refresh_projects(self) -> None:
        self.project_list.delete(0, END)
        self.project_map.clear()
        selected_index: int | None = None
        for row in self.db.list_projects():
            term = f"{row['start_date'] or '-'} - {row['end_date'] or '-'}"
            line = f"{row['name']} / {row['construction_no']} / {row['client_name']} / {term}"
            self.project_map[line] = row["id"]
            current_index = self.project_list.size()
            self.project_list.insert(END, line)
            if row["id"] == self.selection.project_id:
                selected_index = current_index
        if selected_index is not None:
            self.project_list.selection_set(selected_index)
            self.project_list.activate(selected_index)
            self.on_project_selected()
        elif self.project_list.size() > 0 and self.selection.project_id is None:
            self.project_list.selection_set(0)
            self.on_project_selected()
        elif self.project_list.size() == 0:
            self.project_list.insert(END, "工事が登録されていません。「+ 工事登録」または「工事情報Excel取込」から開始してください。")
            self.clear_project_selection()

    def on_project_selected(self, _event: object | None = None) -> None:
        selection = self.project_list.curselection()
        if not selection:
            self.clear_project_selection()
            return
        label = self.project_list.get(selection[0])
        if label not in self.project_map:
            self.clear_project_selection()
            return
        self.selection.project_id = self.project_map[label]
        self.selection.company_id = None
        self.selection.document_id = None
        self.selection.attachment_id = None
        self.selection.relation_document_id = None
        self.selection.relation_attachment_id = None
        self.refresh_company_tree()
        self.refresh_document_tree()
        self.refresh_attachment_tree()
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        self.update_detail(None)

    def clear_project_selection(self) -> None:
        self.selection.project_id = None
        self.selection.company_id = None
        self.selection.document_id = None
        self.selection.attachment_id = None
        self.selection.relation_document_id = None
        self.selection.relation_attachment_id = None
        self.refresh_company_tree()
        self.refresh_document_tree()
        self.refresh_attachment_tree()
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        self.update_detail(None)

    def refresh_company_tree(self) -> None:
        for item in self.company_tree.get_children():
            self.company_tree.delete(item)
        self.company_tree_map.clear()
        selected_tree_id: str | None = None

        project_id = self.selection.project_id
        if project_id is None:
            self.company_tree.insert("", END, text="左の工事を選択してください", values=("", ""))
            return

        companies = self.db.list_companies(project_id)
        if not companies:
            self.company_tree.insert("", END, text="+ 元請 から施工体制を登録してください", values=("", ""))
            return

        by_parent: dict[int | None, list[sqlite3.Row]] = {}
        for row in companies:
            by_parent.setdefault(row["parent_company_id"], []).append(row)

        def insert_nodes(parent_tree_id: str, parent_company_id: int | None) -> None:
            for row in by_parent.get(parent_company_id, []):
                tree_id = self.company_tree.insert(
                    parent_tree_id,
                    END,
                    text=row["name"],
                    values=(LEVEL_LABELS.get(row["level"], f"Lv{row['level']}"), row["work_type"] or ""),
                    open=True,
                )
                self.company_tree_map[tree_id] = row["id"]
                nonlocal selected_tree_id
                if row["id"] == self.selection.company_id:
                    selected_tree_id = tree_id
                insert_nodes(tree_id, row["id"])

        insert_nodes("", None)
        if selected_tree_id is not None:
            self.company_tree.selection_set(selected_tree_id)
            self.company_tree.focus(selected_tree_id)
            self.on_company_selected()

    def on_company_selected(self, _event: object | None = None) -> None:
        selected = self.company_tree.selection()
        if not selected:
            return
        tree_id = selected[0]
        if tree_id not in self.company_tree_map:
            return
        company_id = self.company_tree_map[tree_id]
        self.selection.company_id = company_id
        self.selection.document_id = None
        self.selection.attachment_id = None
        self.selection.relation_attachment_id = None
        relation_document = self.db.get_relation_document_for_child(company_id)
        self.selection.relation_document_id = relation_document["id"] if relation_document is not None else None
        company = self.db.get_company(company_id)
        self.update_detail(company)
        self.refresh_document_tree()
        self.refresh_attachment_tree()
        self.refresh_relation_documents()
        self.refresh_relation_attachments()

    def refresh_document_tree(self) -> None:
        for item in self.document_tree.get_children():
            self.document_tree.delete(item)
        self.document_tree_map.clear()

        company_id = self.selection.company_id
        if company_id is None:
            self.document_tree.insert("", END, values=("業者を選択すると必要添付書類を表示します", "", "", "", "", ""))
            return

        for row in self.db.list_required_documents(company_id):
            attachment_count = int(row["attachment_count"])
            item_id = self.document_tree.insert(
                "",
                END,
                values=(
                    row["document_name"],
                    "○" if row["required"] else "-",
                    row["status"],
                    row["expiry_date"] or "",
                    "0 (未添付)" if attachment_count == 0 else str(attachment_count),
                    row["note"] or "",
                ),
                tags=(row["status"],),
            )
            self.document_tree_map[item_id] = row["id"]

    def on_document_selected(self, _event: object | None = None) -> None:
        selected = self.document_tree.selection()
        if not selected:
            return
        if selected[0] not in self.document_tree_map:
            return
        self.selection.document_id = self.document_tree_map[selected[0]]
        self.selection.attachment_id = None
        self.refresh_attachment_tree()

    def refresh_attachment_tree(self) -> None:
        for item in self.attachment_tree.get_children():
            self.attachment_tree.delete(item)
        self.attachment_tree_map.clear()

        document_id = self.selection.document_id
        if document_id is None:
            self.attachment_tree.insert("", END, values=("書類を選択すると添付ファイルを表示します", "", "", ""))
            return
        for row in self.db.list_attachments(document_id):
            item_id = self.attachment_tree.insert(
                "",
                END,
                values=(row["original_filename"], row["stored_filename"], row["file_type"], row["created_at"]),
            )
            self.attachment_tree_map[item_id] = row["id"]

    def on_attachment_selected(self, _event: object | None = None) -> None:
        selected = self.attachment_tree.selection()
        if not selected:
            return
        if selected[0] not in self.attachment_tree_map:
            return
        self.selection.attachment_id = self.attachment_tree_map[selected[0]]

    def refresh_relation_documents(self) -> None:
        for item in self.relation_document_tree.get_children():
            self.relation_document_tree.delete(item)
        self.relation_document_tree_map.clear()

        project_id = self.selection.project_id
        if project_id is None:
            self.relation_document_tree.insert("", END, values=("工事を選択してください", "", "", "", "", ""))
            self.relation_summary_var.set("工事を選択してください。")
            return

        companies = self.db.list_companies(project_id)
        child_companies = [row for row in companies if row["parent_company_id"] is not None]
        if not child_companies:
            self.relation_document_tree.insert("", END, values=("+ 子業者を登録すると契約関係帳票を管理できます", "", "", "", "", ""))
            self.relation_summary_var.set("子業者を選択すると、この親子関係に必要な帳票を表示します。")
            return

        for child in child_companies:
            self.db.ensure_relation_document_for_child(child["id"])

        selected_tree_id: str | None = None
        selected_relation: sqlite3.Row | None = None
        for row in self.db.list_relation_documents(project_id):
            source_name = Path(row["source_path"]).name if row["source_path"] else ""
            attachment_count = int(row["attachment_count"])
            tree_id = self.relation_document_tree.insert(
                "",
                END,
                values=(
                    f"{row['parent_company_name']} -> {row['child_company_name']}",
                    row["document_kind"],
                    row["status"],
                    source_name,
                    row["updated_at"],
                    "0 (未添付)" if attachment_count == 0 else str(attachment_count),
                ),
                tags=(row["status"],),
            )
            self.relation_document_tree_map[tree_id] = row["id"]
            if row["id"] == self.selection.relation_document_id:
                selected_tree_id = tree_id
                selected_relation = row

        if selected_tree_id is not None:
            self.relation_document_tree.selection_set(selected_tree_id)
            self.relation_document_tree.focus(selected_tree_id)
        if selected_relation is not None:
            self.update_relation_summary(selected_relation)
        else:
            self.relation_summary_var.set("契約関係帳票を選択すると、対象関係と取込状況を表示します。")

    def on_relation_document_selected(self, _event: object | None = None) -> None:
        selected = self.relation_document_tree.selection()
        if not selected:
            return
        if selected[0] not in self.relation_document_tree_map:
            return
        self.selection.relation_document_id = self.relation_document_tree_map[selected[0]]
        self.selection.relation_attachment_id = None
        relation_document = self.db.get_relation_document(self.selection.relation_document_id)
        if relation_document is not None:
            self.update_relation_summary(relation_document)
        self.refresh_relation_attachments()

    def update_relation_summary(self, relation_document: sqlite3.Row) -> None:
        source_name = Path(relation_document["source_path"]).name if relation_document["source_path"] else "未取込"
        self.relation_summary_var.set(
            f"対象関係: {relation_document['parent_company_name']} -> {relation_document['child_company_name']} / "
            f"帳票種別: {relation_document['document_kind']} / "
            f"状態: {relation_document['status']} / "
            f"取込元: {source_name}"
        )

    def refresh_relation_attachments(self) -> None:
        for item in self.relation_attachment_tree.get_children():
            self.relation_attachment_tree.delete(item)
        self.relation_attachment_tree_map.clear()

        relation_document_id = self.selection.relation_document_id
        if relation_document_id is None:
            self.relation_attachment_tree.insert("", END, values=("契約関係帳票を選択すると添付ファイルを表示します", "", "", ""))
            return
        for row in self.db.list_relation_attachments(relation_document_id):
            item_id = self.relation_attachment_tree.insert(
                "",
                END,
                values=(row["original_filename"], row["stored_filename"], row["file_type"], row["created_at"]),
            )
            self.relation_attachment_tree_map[item_id] = row["id"]

    def on_relation_attachment_selected(self, _event: object | None = None) -> None:
        selected = self.relation_attachment_tree.selection()
        if not selected:
            return
        if selected[0] not in self.relation_attachment_tree_map:
            return
        self.selection.relation_attachment_id = self.relation_attachment_tree_map[selected[0]]

    def update_detail(self, company: sqlite3.Row | None) -> None:
        self.detail_text.configure(state="normal")
        self.detail_text.delete("1.0", END)
        if company is None:
            self.detail_text.insert("1.0", "業者を選択すると詳細を表示します。\n\n工事を選択後、施工体制ツリーから対象業者を選択してください。")
        else:
            lines = [
                "基本情報",
                "------------------------------",
                f"会社名: {company['name']}",
                f"階層: {LEVEL_LABELS.get(company['level'], company['level'])}",
                f"会社名カナ: {company['kana'] or ''}",
                f"代表者名: {company['representative'] or ''}",
                f"所在地: {company['address'] or ''}",
                f"電話番号: {company['phone'] or ''}",
                "",
                "許可・担当",
                "------------------------------",
                f"建設業許可番号: {company['license_no'] or ''}",
                f"許可有効期限: {company['license_expiry'] or ''}",
                f"担当工種: {company['work_type'] or ''}",
                "",
                "契約・施工予定",
                "------------------------------",
                f"契約日: {company['contract_date'] or ''}",
                f"施工開始予定日: {company['planned_start_date'] or ''}",
                f"施工終了予定日: {company['planned_end_date'] or ''}",
                "",
                "技術者・安全衛生",
                "------------------------------",
                f"主任技術者名: {company['chief_engineer_name'] or ''}",
                f"主任技術者資格: {company['chief_engineer_license'] or ''}",
                f"安全衛生責任者: {company['safety_manager'] or ''}",
            ]
            self.detail_text.insert("1.0", "\n".join(lines))
        self.detail_text.configure(state="disabled")

    def add_project(self) -> None:
        dialog = ProjectDialog(self.root)
        self.root.wait_window(dialog.top)
        if dialog.result is None:
            return
        project_id = self.db.create_project(dialog.result)
        self.selection.project_id = project_id
        self.refresh_projects()
        self.status_var.set(f"工事を登録しました: {dialog.result['name']}")

    def edit_selected_project(self) -> None:
        project_id = self.selection.project_id
        if project_id is None:
            messagebox.showinfo("未選択", "編集する工事を選択してください。")
            return
        project = self.db.get_project(project_id)
        if project is None:
            return
        initial_data = {
            "name": project["name"] or "",
            "construction_no": project["construction_no"] or "",
            "client_name": project["client_name"] or "",
            "start_date": project["start_date"] or "",
            "end_date": project["end_date"] or "",
            "site_agent": project["site_agent"] or "",
            "managing_engineer": project["managing_engineer"] or "",
            "chief_engineer": project["chief_engineer"] or "",
            "base_folder": project["base_folder"] or "",
        }
        dialog = ProjectDialog(self.root, title="工事編集", initial_data=initial_data)
        self.root.wait_window(dialog.top)
        if dialog.result is None:
            return
        self.db.update_project(project_id, dialog.result)
        self.refresh_projects()
        self.status_var.set(f"工事を更新しました: {dialog.result['name']}")

    def delete_selected_project(self) -> None:
        project_id = self.selection.project_id
        if project_id is None:
            messagebox.showinfo("未選択", "削除する工事を選択してください。")
            return
        project = self.db.get_project(project_id)
        if project is None:
            return
        confirmed = messagebox.askyesno(
            "工事削除",
            f"工事「{project['name']}」を削除します。\n関連する業者、書類、添付情報も削除されます。よろしいですか。",
        )
        if not confirmed:
            return
        self.db.delete_project(project_id)
        self.selection.project_id = None
        self.refresh_projects()
        self.status_var.set(f"工事を削除しました: {project['name']}")

    def reset_debug_data(self) -> None:
        confirmed = messagebox.askyesno(
            "開発用初期化",
            "開発用初期化を実行します。\n登録済みの工事、業者、書類、添付情報、取込履歴をすべて削除します。\n通常運用では使用しない操作です。続行しますか。",
        )
        if not confirmed:
            return
        final_confirmed = messagebox.askyesno(
            "最終確認",
            "本当に全データを初期化しますか。\nこの操作は取り消せません。",
        )
        if not final_confirmed:
            return
        self.db.reset_all_data()
        if ATTACHMENTS_ROOT.exists():
            shutil.rmtree(ATTACHMENTS_ROOT, ignore_errors=True)
        ATTACHMENTS_ROOT.mkdir(parents=True, exist_ok=True)
        self.clear_project_selection()
        self.refresh_projects()
        self.status_var.set("開発用初期化を実行しました。")

    def import_projects_csv(self) -> None:
        csv_path = filedialog.askopenfilename(
            title="工事情報CSVを選択",
            filetypes=[("CSVファイル", "*.csv"), ("すべてのファイル", "*.*")],
        )
        if not csv_path:
            return

        records = self.load_project_csv(Path(csv_path))
        if not records:
            messagebox.showinfo("CSV取込", "取り込める工事データがありませんでした。")
            return

        preview_rows: list[dict[str, str]] = []
        for record in records:
            existing = self.db.find_project_by_construction_no(record["construction_no"])
            preview_rows.append(
                {
                    **record,
                    "action": "更新" if existing else "新規",
                    "project_id": str(existing["id"]) if existing else "",
                }
            )

        dialog = CsvImportConfirmDialog(self.root, preview_rows)
        self.root.wait_window(dialog.top)
        if not dialog.approved:
            return

        imported = 0
        updated = 0
        last_project_id: int | None = None
        for row in preview_rows:
            payload = {
                "name": row["name"],
                "construction_no": row["construction_no"],
                "client_name": row["client_name"],
                "start_date": row["start_date"],
                "end_date": row["end_date"],
                "site_agent": row["site_agent"],
                "managing_engineer": row["managing_engineer"],
                "chief_engineer": row["chief_engineer"],
                "base_folder": row["base_folder"],
            }
            if row["action"] == "更新" and row["project_id"]:
                project_id = int(row["project_id"])
                self.db.update_project(project_id, payload)
                updated += 1
                last_project_id = project_id
            else:
                project_id = self.db.create_project(payload)
                imported += 1
                last_project_id = project_id

        self.selection.project_id = last_project_id
        self.refresh_projects()
        self.status_var.set(f"CSV取込完了: 新規 {imported} 件 / 更新 {updated} 件")

    def import_ledger_excel(self) -> None:
        xlsx_path = filedialog.askopenfilename(
            title="施工体制台帳Excelを選択",
            filetypes=[("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")],
        )
        if not xlsx_path:
            return

        try:
            record = self.load_ledger_project_info(Path(xlsx_path))
        except (FileNotFoundError, ValueError) as exc:
            messagebox.showerror("工事情報Excel取込", str(exc))
            return
        if record is None:
            messagebox.showerror(
                "工事情報Excel取込",
                "施工体制台帳Excelから工事情報を読み取れませんでした。\n先頭シートに工事名称、発注者、工期などの項目があるか確認してください。",
            )
            return

        mapped_items = record.get("mapped_items", [])
        self.last_ledger_import_items = mapped_items if isinstance(mapped_items, list) else []
        existing = self.db.find_project_by_name(record["name"])
        if existing:
            payload = {
                "name": record["name"],
                "construction_no": existing["construction_no"] or provisional_construction_no(record["name"]),
                "client_name": record["client_name"],
                "start_date": record["start_date"],
                "end_date": record["end_date"],
                "site_agent": record["site_agent"],
                "managing_engineer": record["managing_engineer"],
                "chief_engineer": existing["chief_engineer"] or record["chief_engineer"],
                "base_folder": existing["base_folder"] or "",
            }
            self.db.update_project(existing["id"], payload)
            project_id = existing["id"]
            action = "更新"
        else:
            payload = {
                "name": record["name"],
                "construction_no": provisional_construction_no(record["name"]),
                "client_name": record["client_name"],
                "start_date": record["start_date"],
                "end_date": record["end_date"],
                "site_agent": record["site_agent"],
                "managing_engineer": record["managing_engineer"],
                "chief_engineer": record["chief_engineer"],
                "base_folder": "",
            }
            project_id = self.db.create_project(payload)
            action = "新規"

        self.db.upsert_project_import(project_id, "施工体制台帳", str(record["source_path"]), record["source_sheet"])
        self.selection.project_id = project_id
        self.refresh_projects()
        self.status_var.set(f"工事情報Excel取込完了: {action} / {record['name']}")
        self.show_ledger_import_summary(action, record, self.last_ledger_import_items)

    def show_ledger_import_summary(self, action: str, record: dict[str, str], mapped_items: list[dict[str, str]]) -> None:
        filled_items = [item for item in mapped_items if item["field"] != "__sheet_name" and item["value"]]
        preview_lines = []
        for item in filled_items[:12]:
            value = item["value"]
            if len(value) > 40:
                value = value[:37] + "..."
            preview_lines.append(f"- {item['item_name']} / {item['field']} / {item['cell_range']}: {value}")

        preview = "\n".join(preview_lines) if preview_lines else "- 値あり項目なし"
        messagebox.showinfo(
            "工事情報Excel取込",
            (
                f"取込結果: {action}\n"
                f"工事名: {record['name']}\n"
                f"OK行の取得対象: {max(len(mapped_items) - 1, 0)} 件\n"
                f"値あり項目: {len(filled_items)} 件\n\n"
                f"取得値の一部:\n{preview}"
            ),
        )

    def reflect_project_to_ledger(self) -> None:
        project_id = self.selection.project_id
        if project_id is None:
            messagebox.showinfo("未選択", "反映する工事を選択してください。")
            return
        self.export_ledger_from_master()

    def get_selected_relation_document(self) -> sqlite3.Row | None:
        relation_document_id = self.selection.relation_document_id
        if relation_document_id is not None:
            return self.db.get_relation_document(relation_document_id)

        company_id = self.selection.company_id
        if company_id is None:
            return None
        relation_document = self.db.get_relation_document_for_child(company_id)
        if relation_document is not None:
            self.selection.relation_document_id = relation_document["id"]
        return relation_document

    def import_relation_document_excel(self) -> None:
        relation_document = self.get_selected_relation_document()
        company_id = self.selection.company_id
        if relation_document is not None:
            company_id = relation_document["child_company_id"]
        if company_id is None:
            messagebox.showinfo("未選択", "施工体制ツリーで子業者、または契約関係帳票の行を選択してください。")
            return
        child = self.db.get_company(company_id)
        if child is None or child["parent_company_id"] is None:
            messagebox.showinfo("対象外", "元請には親会社がないため、契約関係帳票は子業者を選択して取り込みます。")
            return

        xlsx_path = filedialog.askopenfilename(
            title="契約関係帳票Excelを選択",
            filetypes=[("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")],
        )
        if not xlsx_path:
            return

        workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
        sheet_name = workbook.sheetnames[0]
        workbook.close()
        relation_document_id = self.db.upsert_relation_document_import(company_id, str(Path(xlsx_path)), sheet_name)
        if relation_document_id is None:
            messagebox.showerror("契約関係帳票取込", "親子関係を特定できませんでした。")
            return

        self.selection.relation_document_id = relation_document_id
        relation_document = self.db.get_relation_document(relation_document_id)
        if relation_document is not None and relation_document["document_kind"] == LEDGER_DOCUMENT_KIND:
            try:
                self.apply_ledger_mapped_values_to_db(Path(xlsx_path), relation_document)
            except (FileNotFoundError, ValueError) as exc:
                messagebox.showerror("契約関係帳票取込", str(exc))
                return
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        if relation_document is not None:
            self.status_var.set(f"契約関係帳票を取り込みました: {relation_document['document_kind']} / {Path(xlsx_path).name}")

    def edit_relation_document_info(self) -> None:
        relation_document = self.get_selected_relation_document()
        if relation_document is None:
            messagebox.showinfo("未選択", "編集する契約関係帳票を選択してください。")
            return
        dialog = RelationDocumentDialog(self.root, relation_document)
        self.root.wait_window(dialog.top)
        if dialog.result is None:
            return
        self.db.update_relation_document_note(relation_document["id"], dialog.result["status"], dialog.result["note"])
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        updated = self.db.get_relation_document(relation_document["id"])
        if updated is not None:
            self.update_relation_summary(updated)
        self.status_var.set(f"帳票情報を更新しました: {relation_document['document_kind']}")

    def replace_relation_document_file(self) -> None:
        relation_document = self.get_selected_relation_document()
        if relation_document is None:
            messagebox.showinfo("未選択", "差し替える契約関係帳票を選択してください。")
            return
        xlsx_path = filedialog.askopenfilename(
            title="差し替える契約関係帳票Excelを選択",
            filetypes=[("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")],
        )
        if not xlsx_path:
            return
        workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
        sheet_name = workbook.sheetnames[0]
        workbook.close()
        self.db.update_relation_document_import(relation_document["id"], str(Path(xlsx_path)), sheet_name)
        if relation_document["document_kind"] == LEDGER_DOCUMENT_KIND:
            try:
                self.apply_ledger_mapped_values_to_db(Path(xlsx_path), relation_document)
            except (FileNotFoundError, ValueError) as exc:
                messagebox.showerror("取込元差し替え", str(exc))
                return
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        updated = self.db.get_relation_document(relation_document["id"])
        if updated is not None:
            self.update_relation_summary(updated)
        self.status_var.set(f"取込元ファイルを差し替えました: {Path(xlsx_path).name}")

    def clear_relation_document_registration(self) -> None:
        relation_document = self.get_selected_relation_document()
        if relation_document is None:
            messagebox.showinfo("未選択", "削除する契約関係帳票を選択してください。")
            return
        confirmed = messagebox.askyesno(
            "帳票登録削除",
            "この帳票登録を削除します。\nCIVIL HUB上の紐づけは削除されますが、元のExcel/PDFファイルは削除されません。\nよろしいですか？",
        )
        if not confirmed:
            return
        self.db.clear_relation_document_registration(relation_document["id"])
        self.selection.relation_document_id = relation_document["id"]
        self.selection.relation_attachment_id = None
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        updated = self.db.get_relation_document(relation_document["id"])
        if updated is not None:
            self.update_relation_summary(updated)
        self.status_var.set(f"帳票登録を削除しました: {relation_document['document_kind']}")

    def reflect_relation_document(self) -> None:
        relation_document = self.get_selected_relation_document()
        if relation_document is None:
            messagebox.showinfo("未選択", "施工体制ツリーで子業者、または契約関係帳票一覧の行を選択してください。")
            return
        if relation_document["document_kind"] == LEDGER_DOCUMENT_KIND:
            self.export_ledger_from_master(relation_document=relation_document)
            return
        if not relation_document["source_path"]:
            messagebox.showinfo("未取込", "先に契約関係帳票Excelを取り込んでください。")
            return

        source_path = Path(relation_document["source_path"])
        if not source_path.exists():
            messagebox.showerror(
                "取込元Excelなし",
                f"契約関係帳票の取込元Excelが見つかりません。\n移動または削除されていないか確認してください。\n\n{source_path}",
            )
            return

        output_path = build_reflected_copy_path(source_path)
        confirmed = messagebox.askyesno(
            "契約関係帳票 反映コピー作成",
            f"取込元Excelは変更せず、反映済みコピーを作成します。\n\n対象:\n{relation_document['parent_company_name']} -> {relation_document['child_company_name']}\n\n帳票:\n{relation_document['document_kind']}\n\n出力予定ファイル:\n{output_path}\n\n実行しますか。",
        )
        if not confirmed:
            return

        workbook = load_workbook(source_path)
        sheet_name = relation_document["source_sheet"] or workbook.sheetnames[0]
        if sheet_name not in workbook.sheetnames:
            workbook.close()
            messagebox.showerror("シートなし", f"取込時のシートが見つかりません。\n{sheet_name}")
            return

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        workbook.close()
        self.db.update_relation_document_reflected_path(relation_document["id"], str(output_path))
        self.refresh_relation_documents()
        self.status_var.set(f"契約関係帳票の反映コピーを保存しました: {output_path.name}")
        messagebox.showinfo(
            "契約関係帳票 反映コピー作成",
            f"原本は変更せず、反映済みコピーを保存しました。\n\n出力ファイル:\n{output_path.name}\n\n保存先:\n{output_path.parent}",
        )

    def export_ledger_from_master(self, relation_document: sqlite3.Row | None = None) -> None:
        project_id = self.selection.project_id if relation_document is None else relation_document["project_id"]
        if project_id is None:
            messagebox.showinfo("未選択", "出力する工事を選択してください。")
            return
        project = self.db.get_project(project_id)
        if project is None:
            return
        if not LEDGER_MASTER_PATH.exists():
            messagebox.showerror("マスターなし", f"施工体制台帳マスターが見つかりません。\n{LEDGER_MASTER_PATH}")
            return
        if not LEDGER_CELL_MAPPING_PATH.exists():
            messagebox.showerror("セル対応表なし", f"施工体制台帳セル対応表が見つかりません。\n{LEDGER_CELL_MAPPING_PATH}")
            return

        child_company = None
        parent_company = None
        if relation_document is not None:
            child_company = self.db.get_company(relation_document["child_company_id"])
            parent_company = self.db.get_company(relation_document["parent_company_id"])
        elif self.selection.company_id is not None:
            selected_company = self.db.get_company(self.selection.company_id)
            if selected_company is not None and selected_company["parent_company_id"] is not None:
                child_company = selected_company
                parent_company = self.db.get_company(selected_company["parent_company_id"])

        output_path = build_ledger_export_path(project["name"] or "工事", child_company["name"] if child_company is not None else None)
        confirmed = messagebox.askyesno(
            "施工体制台帳 出力",
            f"施工体制台帳マスターは変更せず、exportsへコピーして出力します。\n\n出力予定ファイル:\n{output_path}\n\n実行しますか。",
        )
        if not confirmed:
            return

        workbook = load_workbook(LEDGER_MASTER_PATH)
        self.write_ledger_mapped_workbook(workbook, project, parent_company, child_company)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        workbook.close()
        if relation_document is not None:
            self.db.update_relation_document_reflected_path(relation_document["id"], str(output_path))
            self.refresh_relation_documents()
        self.status_var.set(f"施工体制台帳を出力しました: {output_path.name}")
        messagebox.showinfo(
            "施工体制台帳 出力",
            f"マスターは変更せず、Hub管理版を保存しました。\n\n出力ファイル:\n{output_path.name}\n\n保存先:\n{output_path.parent}",
        )

    def load_project_csv(self, csv_path: Path) -> list[dict[str, str]]:
        standard_field_map = {
            "工事名": "name",
            "工事番号": "construction_no",
            "発注者": "client_name",
            "工期開始日": "start_date",
            "工期終了日": "end_date",
            "現場代理人": "site_agent",
            "監理技術者": "managing_engineer",
            "主任技術者": "chief_engineer",
            "保存フォルダ": "base_folder",
        }

        rows: list[dict[str, str]] = []
        loaded: list[dict[str, str]] | None = None
        for encoding in ("utf-8-sig", "cp932", "utf-8"):
            try:
                with csv_path.open("r", encoding=encoding, newline="") as handle:
                    reader = csv.DictReader(handle)
                    loaded = list(reader)
                break
            except UnicodeDecodeError:
                continue

        if loaded is None:
            messagebox.showerror("CSV取込", "CSVの文字コードを読み取れませんでした。")
            return rows

        if not loaded:
            return rows

        header_keys = set(loaded[0].keys())
        is_project_config_csv = {"project_name", "excel_output_dir"}.issubset(header_keys)

        if is_project_config_csv:
            for source in loaded:
                project_name = (source.get("project_name", "") or "").strip()
                if not project_name:
                    continue
                output_dir = (source.get("excel_output_dir", "") or "").strip()
                if output_dir:
                    base_folder = str((csv_path.parent / output_dir).resolve()) if not Path(output_dir).is_absolute() else output_dir
                else:
                    base_folder = ""
                rows.append(
                    {
                        "name": project_name,
                        "construction_no": provisional_construction_no(project_name),
                        "client_name": "",
                        "start_date": "",
                        "end_date": "",
                        "site_agent": "",
                        "managing_engineer": "",
                        "chief_engineer": "",
                        "base_folder": base_folder,
                    }
                )
            return rows

        for source in loaded:
            record = {dest: (source.get(src, "") or "").strip() for src, dest in standard_field_map.items()}
            if not record["construction_no"] or not record["name"]:
                continue
            rows.append(record)
        return rows

    def load_ledger_project_info(self, xlsx_path: Path) -> dict[str, str] | None:
        mapped_items = self.read_ledger_mapped_items(xlsx_path)
        values = self.ledger_items_to_values(mapped_items)

        project_name = values.get("project.basic.name", "")
        if not project_name:
            return None

        return {
            "name": project_name,
            "client_name": values.get("client.name", ""),
            "start_date": parse_wareki_free_date(values.get("project.contract.period_start", "").replace("自", "").strip()),
            "end_date": parse_wareki_free_date(values.get("project.contract.period_end", "").replace("至", "").strip()),
            "site_agent": values.get("prime_contractor.engineers.site_agent_name", ""),
            "managing_engineer": values.get("prime_contractor.engineers.chief_engineer_name", ""),
            "chief_engineer": values.get("prime_contractor.engineers.chief_engineer_name", ""),
            "source_path": str(xlsx_path),
            "source_sheet": values.get("__sheet_name", ""),
            "mapped_items": mapped_items,
        }

    def read_ledger_mapped_values(self, xlsx_path: Path) -> dict[str, str]:
        return self.ledger_items_to_values(self.read_ledger_mapped_items(xlsx_path))

    def ledger_items_to_values(self, mapped_items: list[dict[str, str]]) -> dict[str, str]:
        values: dict[str, str] = {}
        for item in mapped_items:
            if item["field"] == "__sheet_name":
                values["__sheet_name"] = item["actual_sheet_name"]
                continue
            if not item["value"]:
                continue
            values[item["field"]] = item["value"]
        return values

    def read_ledger_mapped_items(self, xlsx_path: Path) -> list[dict[str, str]]:
        mappings = load_ledger_cell_mappings()
        workbook = load_workbook(xlsx_path, data_only=True)
        try:
            source_sheet_name = workbook.sheetnames[0]
            source_sheet = workbook[source_sheet_name]
            mapped_items: list[dict[str, str]] = [
                {
                    "sheet_name": source_sheet_name,
                    "actual_sheet_name": source_sheet_name,
                    "cell": "",
                    "cell_range": "",
                    "item_name": "読取シート",
                    "field": "__sheet_name",
                    "value": source_sheet_name,
                }
            ]
            for mapping in mappings:
                sheet_name = mapping["sheet_name"]
                if sheet_name in workbook.sheetnames:
                    actual_sheet_name = sheet_name
                    sheet = workbook[sheet_name]
                else:
                    actual_sheet_name = source_sheet_name
                    sheet = source_sheet
                value = sheet[mapping["cell"]].value
                mapped_items.append(
                    {
                        "sheet_name": mapping["sheet_name"],
                        "actual_sheet_name": actual_sheet_name,
                        "cell": mapping["cell"],
                        "cell_range": mapping["cell_range"],
                        "item_name": mapping["item_name"],
                        "field": mapping["field"],
                        "value": "" if value in (None, "") else str(value).strip(),
                    }
                )
            return mapped_items
        finally:
            workbook.close()

    def write_ledger_mapped_workbook(
        self,
        workbook: object,
        project: sqlite3.Row,
        parent_company: sqlite3.Row | None,
        child_company: sqlite3.Row | None,
    ) -> None:
        for mapping in load_ledger_cell_mappings():
            sheet_name = mapping["sheet_name"]
            if sheet_name not in workbook.sheetnames:
                continue
            value = self.ledger_field_value(mapping["field"], project, parent_company, child_company)
            if value in (None, ""):
                continue
            workbook[sheet_name][mapping["cell"]] = value

    def ledger_field_value(
        self,
        field: str,
        project: sqlite3.Row,
        parent_company: sqlite3.Row | None,
        child_company: sqlite3.Row | None,
    ) -> str:
        root_company = parent_company
        if root_company is None:
            companies = self.db.list_companies(project["id"])
            root_company = next((row for row in companies if int(row["level"]) == 0), None)

        project_values = {
            "project.basic.name": project["name"] or "",
            "selected_contractor.basic.project_name": project["name"] or "",
            "client.name": project["client_name"] or "",
            "project.contract.period_start": format_ledger_date("自", project["start_date"] or ""),
            "project.contract.period_end": format_ledger_date("至", project["end_date"] or ""),
            "prime_contractor.engineers.site_agent_name": project["site_agent"] or "",
            "prime_contractor.engineers.chief_engineer_name": project["managing_engineer"] or project["chief_engineer"] or "",
        }
        if field in project_values:
            return project_values[field]

        if field.startswith("prime_contractor.") and root_company is not None:
            return self.company_ledger_field_value(field, root_company, is_prime=True)
        if field.startswith("selected_contractor.") and child_company is not None:
            return self.company_ledger_field_value(field, child_company, is_prime=False)
        return ""

    def company_ledger_field_value(self, field: str, company: sqlite3.Row, is_prime: bool) -> str:
        if field.endswith(".company_name_with_corporate_no") or field.endswith(".office_name"):
            return company["name"] or ""
        if field.endswith(".representative_name"):
            return company["representative"] or ""
        if field.endswith(".address"):
            return company["address"] or ""
        if field.endswith(".phone"):
            return format_ledger_phone(company["phone"] or "")
        if field.endswith(".work_description") or field.endswith(".work_description_1") or field.endswith(".work_description_2"):
            return company["work_type"] or ""
        if field.endswith(".period_start"):
            return format_ledger_date("自", company["planned_start_date"] or "")
        if field.endswith(".period_end"):
            return format_ledger_date("至", company["planned_end_date"] or "")
        if field.endswith(".contract_date"):
            return format_ledger_date("", company["contract_date"] or "").strip()
        if field.endswith(".permit_business_type_1"):
            return company["work_type"] or ""
        if field.endswith(".permit_number_1"):
            return company["license_no"] or ""
        if field.endswith(".permit_date_1"):
            return format_ledger_date("", company["license_expiry"] or "").strip()
        if field.endswith(".safety_health_manager_name"):
            return company["safety_manager"] or ""
        if field.endswith(".chief_engineer_name"):
            return company["chief_engineer_name"] or ""
        if field.endswith(".chief_engineer_qualification"):
            return company["chief_engineer_license"] or ""
        if field.endswith(".chief_engineer_assignment_type"):
            return "専任" if company["chief_engineer_name"] else ""
        if is_prime and field.endswith(".site_agent_name"):
            return company["representative"] or ""
        return ""

    def apply_ledger_mapped_values_to_db(self, xlsx_path: Path, relation_document: sqlite3.Row) -> None:
        values = self.read_ledger_mapped_values(xlsx_path)
        project = self.db.get_project(relation_document["project_id"])
        child_company = self.db.get_company(relation_document["child_company_id"])
        if project is not None:
            project_payload = {
                "name": values.get("project.basic.name", project["name"] or ""),
                "construction_no": project["construction_no"] or provisional_construction_no(values.get("project.basic.name", project["name"] or "")),
                "client_name": values.get("client.name", project["client_name"] or ""),
                "start_date": parse_wareki_free_date(values.get("project.contract.period_start", project["start_date"] or "").replace("自", "").strip()),
                "end_date": parse_wareki_free_date(values.get("project.contract.period_end", project["end_date"] or "").replace("至", "").strip()),
                "site_agent": values.get("prime_contractor.engineers.site_agent_name", project["site_agent"] or ""),
                "managing_engineer": values.get("prime_contractor.engineers.chief_engineer_name", project["managing_engineer"] or ""),
                "chief_engineer": project["chief_engineer"] or "",
                "base_folder": project["base_folder"] or "",
            }
            self.db.update_project(project["id"], project_payload)

        if child_company is not None:
            company_payload = {
                "name": values.get("selected_contractor.basic.company_name_with_corporate_no", child_company["name"] or ""),
                "kana": child_company["kana"] or "",
                "representative": values.get("selected_contractor.basic.representative_name", child_company["representative"] or ""),
                "address": clean_ledger_address(values.get("selected_contractor.basic.address", child_company["address"] or "")),
                "phone": clean_ledger_phone(values.get("selected_contractor.basic.phone", child_company["phone"] or "")),
                "license_no": values.get("selected_contractor.permits.permit_number_1", child_company["license_no"] or ""),
                "license_expiry": parse_wareki_free_date(values.get("selected_contractor.permits.permit_date_1", child_company["license_expiry"] or "")),
                "work_type": values.get("selected_contractor.basic.work_description", values.get("selected_contractor.permits.permit_business_type_1", child_company["work_type"] or "")),
                "contract_date": parse_wareki_free_date(values.get("selected_contractor.basic.contract_date", child_company["contract_date"] or "")),
                "planned_start_date": parse_wareki_free_date(values.get("selected_contractor.basic.period_start", child_company["planned_start_date"] or "").replace("自", "").strip()),
                "planned_end_date": parse_wareki_free_date(values.get("selected_contractor.basic.period_end", child_company["planned_end_date"] or "").replace("至", "").strip()),
                "chief_engineer_name": values.get("selected_contractor.engineers.chief_engineer_name", child_company["chief_engineer_name"] or ""),
                "chief_engineer_license": values.get("selected_contractor.engineers.chief_engineer_qualification", child_company["chief_engineer_license"] or ""),
                "safety_manager": values.get("selected_contractor.engineers.safety_health_manager_name", child_company["safety_manager"] or ""),
            }
            self.db.update_company(child_company["id"], company_payload)

    def write_project_to_ledger_sheet(self, sheet: object, project: sqlite3.Row) -> None:
        name = project["name"] or ""
        client_name = project["client_name"] or ""
        start_date = project["start_date"] or ""
        end_date = project["end_date"] or ""
        site_agent = project["site_agent"] or ""
        managing_engineer = project["managing_engineer"] or ""

        # 左側ブロックの主要な工事情報に限定して反映する
        if name:
            sheet["I6"] = name
            sheet["G21"] = name
        if client_name:
            sheet["G25"] = client_name
        if start_date:
            sheet["H29"] = format_ledger_date("自", start_date)
        if end_date:
            sheet["H30"] = format_ledger_date("至", end_date)
        if site_agent:
            sheet["G55"] = site_agent
        if managing_engineer:
            sheet["J57"] = managing_engineer

    def find_sheet_value(
        self,
        sheet: object,
        candidate_labels: list[str],
        offset_rows: int = 0,
        occurrence: int = 0,
        value_index: int = 0,
    ) -> str:
        normalized_labels = [normalize_label(label) for label in candidate_labels]
        matches: list[tuple[int, int]] = []
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                normalized = normalize_label(cell.value)
                if any(label in normalized for label in normalized_labels):
                    matches.append((cell.row, cell.column))

        if not matches or occurrence >= len(matches):
            return ""

        row_index, col_index = matches[occurrence]
        target_row = row_index + offset_rows
        found_values = 0
        for next_col in range(col_index + 1, sheet.max_column + 1):
            value = sheet.cell(target_row, next_col).value
            if value not in (None, ""):
                if found_values == value_index:
                    return str(value).strip()
                found_values += 1
        return ""

    def add_root_company(self) -> None:
        project_id = self.selection.project_id
        if project_id is None:
            messagebox.showinfo("未選択", "先に工事を選択してください。")
            return
        dialog = CompanyDialog(self.root, level=0, parent_name=None)
        self.root.wait_window(dialog.top)
        if dialog.result is None:
            return
        payload: dict[str, str | int | None] = dict(dialog.result)
        payload.update({"project_id": project_id, "parent_company_id": None, "level": 0})
        company_id = self.db.create_company(payload)
        self.selection.company_id = company_id
        self.refresh_company_tree()
        self.status_var.set(f"元請を登録しました: {dialog.result['name']}")

    def add_child_company(self) -> None:
        project_id = self.selection.project_id
        parent_company_id = self.selection.company_id
        if project_id is None or parent_company_id is None:
            messagebox.showinfo("未選択", "先に工事と親業者を選択してください。")
            return
        parent = self.db.get_company(parent_company_id)
        if parent is None:
            return
        level = int(parent["level"]) + 1
        dialog = CompanyDialog(self.root, level=level, parent_name=parent["name"])
        self.root.wait_window(dialog.top)
        if dialog.result is None:
            return
        payload: dict[str, str | int | None] = dict(dialog.result)
        payload.update({"project_id": project_id, "parent_company_id": parent_company_id, "level": level})
        company_id = self.db.create_company(payload)
        self.selection.company_id = company_id
        self.refresh_company_tree()
        self.status_var.set(f"子業者を登録しました: {dialog.result['name']}")

    def edit_selected_company(self) -> None:
        company_id = self.selection.company_id
        if company_id is None:
            messagebox.showinfo("未選択", "編集する会社を施工体制ツリーで選択してください。")
            return
        company = self.db.get_company(company_id)
        if company is None:
            return
        parent_name = None
        if company["parent_company_id"] is not None:
            parent = self.db.get_company(company["parent_company_id"])
            parent_name = parent["name"] if parent is not None else None
        initial_data = {
            "name": company["name"] or "",
            "kana": company["kana"] or "",
            "representative": company["representative"] or "",
            "address": company["address"] or "",
            "phone": company["phone"] or "",
            "license_no": company["license_no"] or "",
            "license_expiry": company["license_expiry"] or "",
            "work_type": company["work_type"] or "",
            "contract_date": company["contract_date"] or "",
            "planned_start_date": company["planned_start_date"] or "",
            "planned_end_date": company["planned_end_date"] or "",
            "chief_engineer_name": company["chief_engineer_name"] or "",
            "chief_engineer_license": company["chief_engineer_license"] or "",
            "safety_manager": company["safety_manager"] or "",
        }
        dialog = CompanyDialog(
            self.root,
            level=int(company["level"]),
            parent_name=parent_name,
            title="会社情報編集",
            initial_data=initial_data,
        )
        self.root.wait_window(dialog.top)
        if dialog.result is None:
            return
        self.db.update_company(company_id, dialog.result)
        self.selection.company_id = company_id
        updated_company = self.db.get_company(company_id)
        self.refresh_company_tree()
        self.refresh_document_tree()
        self.refresh_attachment_tree()
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        self.update_detail(updated_company)
        self.status_var.set(f"会社情報を更新しました: {dialog.result['name']}")

    def edit_selected_document(self) -> None:
        document_id = self.selection.document_id
        if document_id is None:
            messagebox.showinfo("未選択", "更新する書類を選択してください。")
            return
        document = self.db.fetchone("SELECT * FROM required_documents WHERE id = ?", (document_id,))
        if document is None:
            return
        dialog = DocumentStatusDialog(self.root, document)
        self.root.wait_window(dialog.top)
        if dialog.result is None:
            return
        self.db.update_required_document(
            document_id,
            dialog.result["status"],
            dialog.result["expiry_date"],
            dialog.result["note"],
        )
        self.refresh_document_tree()
        self.status_var.set(f"書類状態を更新しました: {document['document_name']}")

    def show_shortages(self) -> None:
        project_id = self.selection.project_id
        if project_id is None:
            messagebox.showinfo("未選択", "工事を選択してください。")
            return
        for company in self.db.list_companies(project_id):
            if company["parent_company_id"] is not None:
                self.db.ensure_relation_document_for_child(company["id"])
        rows = self.db.list_shortages(project_id)
        if not rows:
            messagebox.showinfo("不足一覧", "不足・未確認・期限切れの書類はありません。")
            return
        lines = [f"{index}. [{row['status']}] {row['company_name']} / {row['document_name']}" for index, row in enumerate(rows, start=1)]
        messagebox.showinfo("不足一覧", f"不足・未確認・期限切れ: {len(rows)}件\n\n" + "\n".join(lines))

    def attach_file(self) -> None:
        document_id = self.selection.document_id
        company_id = self.selection.company_id
        project_id = self.selection.project_id
        missing = []
        if project_id is None:
            missing.append("工事")
        if company_id is None:
            missing.append("業者")
        if document_id is None:
            missing.append("必要添付書類")
        if missing:
            messagebox.showinfo("未選択", "添付前に次を選択してください。\n" + "、".join(missing))
            return

        path_str = filedialog.askopenfilename(
            title="添付ファイルを選択",
            filetypes=[
                ("許可ファイル", "*.pdf *.xlsx *.xls *.docx *.doc *.png *.jpg *.jpeg *.bmp"),
                ("すべてのファイル", "*.*"),
            ],
        )
        if not path_str:
            return

        source = Path(path_str)
        project = self.db.get_project(project_id)
        company = self.db.get_company(company_id)
        document = self.db.fetchone("SELECT * FROM required_documents WHERE id = ?", (document_id,))
        if project is None or company is None or document is None:
            return

        project_folder = Path(project["base_folder"]).expanduser() if project["base_folder"] else ATTACHMENTS_ROOT / f"project_{project_id}"
        target_dir = project_folder / "CIVIL_HUB_添付" / sanitize_filename(company["name"]) / sanitize_filename(document["document_name"])
        target_dir.mkdir(parents=True, exist_ok=True)

        target_name = self.db.next_versioned_filename(target_dir, company["name"], document["document_name"], source.suffix.lower())
        target = target_dir / target_name
        shutil.copy2(source, target)
        self.db.add_attachment(document_id, source, target)
        self.db.update_required_document(document_id, "添付済み", document["expiry_date"] or "", document["note"] or "")
        self.refresh_document_tree()
        self.refresh_attachment_tree()
        self.status_var.set(f"ファイルを添付しました: {target.name}")

    def open_attachment(self) -> None:
        attachment_id = self.selection.attachment_id
        if attachment_id is None:
            messagebox.showinfo("未選択", "開く添付ファイルを選択してください。")
            return
        attachment = self.db.get_attachment(attachment_id)
        if attachment is None:
            return
        target = Path(attachment["stored_path"])
        if not target.exists():
            messagebox.showerror("ファイルなし", f"保存ファイルが見つかりません。\n{target}")
            return
        open_path(target)

    def open_attachment_folder(self) -> None:
        attachment_id = self.selection.attachment_id
        if attachment_id is not None:
            attachment = self.db.get_attachment(attachment_id)
            if attachment is not None:
                folder = Path(attachment["stored_path"]).parent
                folder.mkdir(parents=True, exist_ok=True)
                open_path(folder)
                return

        document_id = self.selection.document_id
        company_id = self.selection.company_id
        project_id = self.selection.project_id
        if document_id is None or company_id is None or project_id is None:
            messagebox.showinfo("未選択", "保存場所を開くには、先に工事・業者・必要添付書類を選択してください。")
            return

        project = self.db.get_project(project_id)
        company = self.db.get_company(company_id)
        document = self.db.fetchone("SELECT * FROM required_documents WHERE id = ?", (document_id,))
        if project is None or company is None or document is None:
            return
        project_folder = Path(project["base_folder"]).expanduser() if project["base_folder"] else ATTACHMENTS_ROOT / f"project_{project_id}"
        folder = project_folder / "CIVIL_HUB_添付" / sanitize_filename(company["name"]) / sanitize_filename(document["document_name"])
        folder.mkdir(parents=True, exist_ok=True)
        open_path(folder)

    def relation_document_folder(self, relation_document: sqlite3.Row) -> Path | None:
        project = self.db.get_project(relation_document["project_id"])
        if project is None:
            return None
        project_folder = Path(project["base_folder"]).expanduser() if project["base_folder"] else ATTACHMENTS_ROOT / f"project_{relation_document['project_id']}"
        relation_name = f"{sanitize_filename(relation_document['parent_company_name'])}__{sanitize_filename(relation_document['child_company_name'])}"
        return project_folder / "CIVIL_HUB_関係帳票" / relation_name / sanitize_filename(relation_document["document_kind"])

    def attach_relation_document_file(self) -> None:
        relation_document = self.get_selected_relation_document()
        if relation_document is None:
            messagebox.showinfo("未選択", "施工体制ツリーで子業者、または契約関係帳票一覧の行を選択してください。")
            return

        path_str = filedialog.askopenfilename(
            title="契約関係帳票の添付ファイルを選択",
            filetypes=[
                ("許可ファイル", "*.pdf *.xlsx *.xls *.docx *.doc *.png *.jpg *.jpeg *.bmp"),
                ("すべてのファイル", "*.*"),
            ],
        )
        if not path_str:
            return

        target_dir = self.relation_document_folder(relation_document)
        if target_dir is None:
            return
        target_dir.mkdir(parents=True, exist_ok=True)
        source = Path(path_str)
        target_name = self.db.next_versioned_filename(
            target_dir,
            f"{relation_document['parent_company_name']}_{relation_document['child_company_name']}",
            relation_document["document_kind"],
            source.suffix.lower(),
        )
        target = target_dir / target_name
        shutil.copy2(source, target)
        self.db.add_relation_attachment(relation_document["id"], source, target)
        self.db.update_relation_document_status(relation_document["id"], "添付済み")
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        self.status_var.set(f"契約関係帳票ファイルを添付しました: {target.name}")

    def open_relation_attachment(self) -> None:
        attachment_id = self.selection.relation_attachment_id
        if attachment_id is None:
            messagebox.showinfo("未選択", "開く契約関係帳票添付ファイルを選択してください。")
            return
        attachment = self.db.get_relation_attachment(attachment_id)
        if attachment is None:
            return
        target = Path(attachment["stored_path"])
        if not target.exists():
            messagebox.showerror("ファイルなし", f"保存ファイルが見つかりません。\n{target}")
            return
        open_path(target)

    def delete_relation_attachment_registration(self) -> None:
        attachment_id = self.selection.relation_attachment_id
        if attachment_id is None:
            messagebox.showinfo("未選択", "削除する契約関係帳票の添付登録を選択してください。")
            return
        attachment = self.db.get_relation_attachment(attachment_id)
        if attachment is None:
            return
        confirmed = messagebox.askyesno(
            "帳票添付登録削除",
            f"この添付登録を削除します。\nCIVIL HUB上の紐づけは削除されますが、保存済みファイルは削除されません。\n\n{attachment['stored_filename']}\n\nよろしいですか？",
        )
        if not confirmed:
            return
        self.db.delete_relation_attachment(attachment_id)
        self.selection.relation_attachment_id = None
        self.refresh_relation_documents()
        self.refresh_relation_attachments()
        self.status_var.set(f"帳票添付登録を削除しました: {attachment['stored_filename']}")

    def open_relation_attachment_folder(self) -> None:
        attachment_id = self.selection.relation_attachment_id
        if attachment_id is not None:
            attachment = self.db.get_relation_attachment(attachment_id)
            if attachment is not None:
                folder = Path(attachment["stored_path"]).parent
                folder.mkdir(parents=True, exist_ok=True)
                open_path(folder)
                return

        relation_document = self.get_selected_relation_document()
        if relation_document is None:
            messagebox.showinfo("未選択", "保存場所を開くには、契約関係帳票を選択してください。")
            return
        folder = self.relation_document_folder(relation_document)
        if folder is None:
            return
        folder.mkdir(parents=True, exist_ok=True)
        open_path(folder)

    def show_rename_candidate(self) -> None:
        document_id = self.selection.document_id
        company_id = self.selection.company_id
        project_id = self.selection.project_id
        if document_id is None or company_id is None or project_id is None:
            messagebox.showinfo("未選択", "先に工事・業者・書類を選択してください。")
            return
        project = self.db.get_project(project_id)
        company = self.db.get_company(company_id)
        document = self.db.fetchone("SELECT * FROM required_documents WHERE id = ?", (document_id,))
        if project is None or company is None or document is None:
            return

        suffix = simpledialog.askstring("拡張子", "拡張子を入力してください。例: .pdf", initialvalue=".pdf")
        if not suffix:
            return
        if not suffix.startswith("."):
            suffix = f".{suffix}"

        project_folder = Path(project["base_folder"]).expanduser() if project["base_folder"] else ATTACHMENTS_ROOT / f"project_{project_id}"
        target_dir = project_folder / "CIVIL_HUB_添付" / sanitize_filename(company["name"]) / sanitize_filename(document["document_name"])
        target_dir.mkdir(parents=True, exist_ok=True)
        candidate = self.db.next_versioned_filename(target_dir, company["name"], document["document_name"], suffix.lower())
        messagebox.showinfo("リネーム候補", candidate)


def main() -> None:
    ATTACHMENTS_ROOT.mkdir(parents=True, exist_ok=True)
    root = Tk()
    app = CivilHubApp(root)
    root.minsize(1200, 760)
    root.mainloop()


if __name__ == "__main__":
    main()
